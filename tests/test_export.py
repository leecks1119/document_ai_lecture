import io

from openpyxl import load_workbook

from src.export import receipt_to_rows, receipt_to_xlsx_bytes, safe_spreadsheet_text
from src.sample_data import SAMPLE_RECEIPT


def test_receipt_items_become_excel_rows():
    rows = receipt_to_rows(SAMPLE_RECEIPT)

    assert len(rows) == 2
    assert rows[0]["item_name"] == "연필"


def test_xlsx_has_expected_sheets_and_values():
    xlsx_bytes = receipt_to_xlsx_bytes(
        SAMPLE_RECEIPT,
        source_text="샘플 OCR 원문",
    )
    workbook = load_workbook(io.BytesIO(xlsx_bytes))

    assert workbook.sheetnames == ["검토_요약", "품목", "원문_근거"]
    assert workbook["품목"]["A2"].value == "샘플문구점"
    evidence = {
        row[0].value: row[1].value
        for row in workbook["원문_근거"].iter_rows(min_col=1, max_col=2)
    }
    assert evidence["ocr_text"] == "샘플 OCR 원문"


def test_formula_prefix_is_escaped():
    assert safe_spreadsheet_text("=1+1") == "'=1+1"
    assert safe_spreadsheet_text(" \t=1+1") == "' \t=1+1"
    assert safe_spreadsheet_text("\r@SUM(A1:A2)") == "'\r@SUM(A1:A2)"


def test_formula_prefix_is_escaped_in_xlsx():
    data = {
        **SAMPLE_RECEIPT,
        "store_name": "=HYPERLINK(\"https://example.com\")",
    }
    workbook = load_workbook(io.BytesIO(receipt_to_xlsx_bytes(data)))

    assert workbook["검토_요약"]["D2"].value.startswith("'=")


def test_review_record_is_written():
    review = {
        "decision": "CHANGED",
        "reviewer": "learner-03",
        "reviewed_at": "2026-07-28T14:10:00+09:00",
        "note": "원본 대조 후 합계 수정",
    }
    workbook = load_workbook(
        io.BytesIO(receipt_to_xlsx_bytes(SAMPLE_RECEIPT, review_record=review))
    )

    assert workbook["검토_요약"]["E2"].value == "CHANGED"
    assert workbook["검토_요약"]["F2"].value == "learner-03"
