"""같은 한국 영수증을 1~7교시까지 이어 쓰는 Colab 노트북 생성기."""

from __future__ import annotations

import json
from pathlib import Path
from pprint import pformat
from textwrap import dedent, wrap

ROOT = Path(__file__).resolve().parents[1]
COLAB_DIR = ROOT / "colab"

GOLDEN_OCR_TEXT = """이태리집
거래일시 2025-10-04 12:33:37
페퍼로니 앤 치즈 29,000 1 29,000
토마토 파스타 14,000 1 14,000
수제 돈가스 13,000 1 13,000
새우 칠리치 필라 14,000 1 14,000
콜라 2,000 3 6,000
합계 금액 76,000
부가세 과세물품가액 69,094
부가세 6,906
"""

GOLDEN_VLM_MARKDOWN = """# 이태리집

> **수업용 VLM 구조 예제** — 지금 모델을 실행해 만든 결과가 아닙니다.

거래일시: 2025-10-04 12:33:37

| 품목 | 수량 | 단가 | 금액 |
| --- | ---: | ---: | ---: |
| 페퍼로니 앤 치즈 | 1 | 29,000원 | 29,000원 |
| 토마토 파스타 | 1 | 14,000원 | 14,000원 |
| 수제 돈가스 | 1 | 13,000원 | 13,000원 |
| 새우 칠리치 필라 | 1 | 14,000원 | 14,000원 |
| 콜라 | 3 | 2,000원 | 6,000원 |

**합계: 76,000원**

부가세 과세물품가액 69,094
부가세 6,906
"""

GOLDEN_RECEIPT = {
    "document_type": "receipt",
    "store_name": "이태리집",
    "date": "2025-10-04",
    "total_amount": 76000,
    "items": [
        {
            "name": "페퍼로니 앤 치즈",
            "quantity": 1,
            "unit_price": 29000,
            "line_total": 29000,
        },
        {
            "name": "토마토 파스타",
            "quantity": 1,
            "unit_price": 14000,
            "line_total": 14000,
        },
        {
            "name": "수제 돈가스",
            "quantity": 1,
            "unit_price": 13000,
            "line_total": 13000,
        },
        {
            "name": "새우 칠리치 필라",
            "quantity": 1,
            "unit_price": 14000,
            "line_total": 14000,
        },
        {
            "name": "콜라",
            "quantity": 3,
            "unit_price": 2000,
            "line_total": 6000,
        },
    ],
    "adjustments": {"discount": 0, "tax": 0, "service": 0, "rounding": 0},
    "tax_breakdown": {
        "mode": "included_in_item_prices",
        "supply_amount": 69094,
        "vat": 6906,
        "payable_total": 76000,
    },
    "raw_values": {
        "store_name": "이태리집",
        "date": "2025-10-04 12:33:37",
        "total_amount": "76,000",
    },
    "cleaned_values": {
        "store_name": "이태리집",
        "date": "2025-10-04",
        "total_amount": 76000,
    },
    "evidence": {
        "store_name": {"raw_value": "이태리집", "line": 1},
        "date": {"raw_value": "거래일시 2025-10-04 12:33:37", "line": 2},
        "total_amount": {"raw_value": "합계 금액 76,000", "line": 8},
    },
    "source_mode": "course_example_rule_extraction",
}

EXTENSION_EXAMPLES = {
    "quotation": {
        "name": "견적서",
        "fields": ["문서번호", "공급자", "수신", "견적일", "품목", "총액"],
        "rules": ["수량×단가=품목금액", "공급가액+부가세=총액"],
        "risk": "총액 오류는 구매 의사결정에 직접 영향",
    },
    "application": {
        "name": "신청서",
        "fields": ["신청번호", "신청자", "소속", "신청 과정", "승인"],
        "rules": ["필수 동의", "관리자 승인 상태"],
        "risk": "개인정보와 승인 누락을 사람이 확인",
    },
    "transaction_statement": {
        "name": "거래명세서",
        "fields": ["문서번호", "공급자", "거래일", "품목", "세액", "총액"],
        "rules": ["품목 합계=공급가액", "공급가액+세액=총액"],
        "risk": "표 행·열 대응이 어긋나면 정산 오류",
    },
}

NOTEBOOK_SLUGS = {
    1: "document_ai_overview",
    2: "ocr_basic",
    3: "document_structure",
    4: "genai_extraction",
    5: "streamlit_basic",
    6: "ocr_ai_integration",
    7: "validation_export",
    8: "business_application",
}

LESSON_BEGINNER_PATHS = {
    "01_document_ai_overview.ipynb": {
        "required": "제공된 한국 영수증 한 장에서 OCR·VLM·Document AI 결과가 어떻게 다른지 비교합니다.",
        "edit": "원본 관찰값과 마지막 기술 선택만 필수입니다. 중간 판단·수정은 공개 정답을 보며 따라가도 됩니다.",
        "research": "2교시에서 다른 공개 영수증이나 비식별 영수증으로 OCR을 다시 실행합니다.",
    },
    "02_ocr_basic.ipynb": {
        "required": "제공된 한국 영수증을 PaddleOCR로 직접 읽고 상자가 글자 위에 맞는지 확인합니다.",
        "edit": "`USE_MY_FILE` 한 곳만 `True`로 바꾸면 내 사진이나 인터넷에서 내려받은 공개 이미지도 시험할 수 있습니다.",
        "research": "문서 한 장만 바꾸어 사진 기울기·해상도·언어에 따라 박스와 글자가 어떻게 달라지는지 기록합니다.",
    },
    "03_document_structure.ipynb": {
        "required": "2교시 OCR 결과의 낱말 좌표를 사람이 읽는 행과 문서 영역으로 다시 묶습니다.",
        "edit": "품목 행을 찾는 정규식 한 줄만 채웁니다. 정답을 복사해 실행해도 됩니다.",
        "research": "2교시에서 만든 다른 이미지의 `ocr_result.json`을 넣어 행 묶기가 어디서 깨지는지 비교합니다.",
    },
    "04_genai_extraction.ipynb": {
        "required": "OCR 원문에서 날짜·합계·품목과 각 값의 원문 근거를 JSON으로 만듭니다.",
        "edit": "Excel 저장 전에 확인할 검토 결정 세 곳만 채웁니다.",
        "research": "다른 OCR 결과를 넣어 규칙 추출과 VLM 구조 예제가 놓치는 필드를 비교합니다.",
    },
    "05_streamlit_basic.ipynb": {
        "required": "지금까지 만든 처리 결과를 파일 업로드·실행 버튼·결과 영역이 있는 웹앱으로 감쌉니다.",
        "edit": "앱 제목과 버튼 문구 두 곳만 업무에 맞게 바꿉니다.",
        "research": "내가 고른 문서를 처음 보는 동료도 버튼의 행동을 이해할 수 있는지 확인합니다.",
    },
    "06_ocr_ai_integration.ipynb": {
        "required": "웹앱에 문서 한 장을 올리고 현재 파일을 OCR로 읽어 구조화 결과까지 연결합니다.",
        "edit": "업무에서 반드시 맞아야 할 필드 세 곳만 고릅니다.",
        "research": "2교시에서 고른 공개 이미지를 앱에 넣어 성공·실패 이유와 필수 필드 결과를 비교합니다.",
    },
    "07_validation_export.ipynb": {
        "required": "잘못된 값은 저장을 막고, 원본 확인과 승인 뒤에만 Excel을 내려받습니다.",
        "edit": "승인 여부·검토자·원본 확인 메모 세 곳만 채웁니다.",
        "research": "내 실험 문서에서 틀린 값 하나를 일부러 넣어 어떤 규칙이 저장을 막아야 하는지 확인합니다.",
    },
    "08_business_application.ipynb": {
        "required": "견적서·신청서·거래명세서 중 한 문서를 골라 작은 PoC 후보 카드를 만듭니다.",
        "edit": "문서 종류·점수·검토자·중단 조건만 채웁니다.",
        "research": "공개 PDF나 비식별 캡처 한 장을 찾아 필요한 필드와 실패 조건을 카드에 추가합니다.",
    },
}

LESSON_STEP_GUIDES = {
    "01_document_ai_overview.ipynb": [
        (
            "실습 환경 준비",
            "필요한 라이브러리와 결과 저장 폴더를 준비합니다.",
            "`실습 준비 완료` 문구가 보이는지 확인합니다.",
        ),
        (
            "실습 자료 불러오기 기능 준비",
            "저장소 다운로드와 수동 업로드 복구 기능을 등록합니다.",
            "오류 없이 셀이 끝나면 자료 로더가 준비된 것입니다.",
        ),
        (
            "원본 영수증 확인",
            "공개 영수증과 실제 PP-OCRv5 기록을 불러옵니다.",
            "영수증 이미지와 `이미지 출처`가 표시되어야 합니다.",
        ),
        (
            "사람이 먼저 읽기",
            "원본에서 상호명·날짜·품목 수·합계를 직접 입력합니다.",
            "내가 입력한 네 값이 한 행의 표로 보여야 합니다.",
        ),
        (
            "원본 관찰 정답 비교",
            "내 관찰값을 공개 정답과 나란히 비교합니다.",
            "네 항목별로 일치·불일치·미작성 상태를 확인합니다.",
        ),
        (
            "OCR 결과 관찰",
            "OCR 글자 위치와 신뢰도를 원본 이미지 위에서 확인합니다.",
            "색상 사각형 이미지와 핵심 판독값 네 개가 보여야 합니다.",
        ),
        (
            "OCR 결과 판단",
            "원본과 OCR 판독값을 비교해 다섯 질문에 답합니다.",
            "내 답과 전체 정답의 비교표를 확인합니다.",
        ),
        (
            "VLM 구조 초안 관찰",
            "품목 표로 정리된 구조 초안과 일부러 넣은 오류를 찾습니다.",
            "품목 합계 76,000원과 초안 합계 16,000원이 보여야 합니다.",
        ),
        (
            "VLM 초안 판단",
            "구조화 성공과 값의 정확성을 나누어 판단합니다.",
            "다섯 질문의 내 답·전체 정답 비교표를 확인합니다.",
        ),
        (
            "Document AI 검증 실행",
            "합계·원본 확인값·원문 근거 규칙으로 초안을 검사합니다.",
            "저장을 막는 오류 세 개가 표로 표시되어야 합니다.",
        ),
        (
            "내 수정값 재검증",
            "합계와 원문 근거를 입력하고 검증 함수를 다시 실행합니다.",
            "검증 통과 여부와 남은 문제가 표시되어야 합니다.",
        ),
        (
            "수정 정답 확인",
            "공개 정답으로 오류가 사라지는지 확인합니다.",
            "수정 후 76,000원, 원문 근거, `통과`가 보여야 합니다.",
        ),
        (
            "사람 승인 구분",
            "자동 검증 통과와 사람 승인이 별도 조건임을 확인합니다.",
            "승인 전 저장 대기와 승인 후 결과표를 차례로 확인합니다.",
        ),
        (
            "기술 역할 구분",
            "OCR·멀티모달 AI·VLM·Document AI·IDP를 상황과 연결합니다.",
            "다섯 상황의 내 답과 전체 정답 비교표를 확인합니다.",
        ),
        (
            "1교시 결과 저장",
            "관찰·수정·승인·개념 답변을 하나의 JSON으로 저장합니다.",
            "`CHECKPOINT 1/1 PASS`와 결과 파일 경로를 확인합니다.",
        ),
    ],
    "02_ocr_basic.ipynb": [
        (
            "공통 환경 준비",
            "결과 폴더와 실습 자료 다운로드 기능을 준비합니다.",
            "Python·Platform·공통 작업 폴더가 표시되어야 합니다.",
        ),
        (
            "OCR 입력 한 장 준비",
            "공개 영수증과 장애 시 사용할 검수 데이터를 불러옵니다.",
            "직접 읽을 파일명과 처리 계획이 표시되어야 합니다.",
        ),
        (
            "PP-OCRv5 실행",
            "영수증 한 장을 지금 OCR로 읽고, 실패하면 예제 결과임을 분명히 표시합니다.",
            "처리 방식·실패 이유·판독 영역 수를 확인합니다.",
        ),
        (
            "OCR 위치 시각화와 저장",
            "판독 영역을 원본 위에 그리고 JSON과 이미지를 저장합니다.",
            "`CHECKPOINT 1/1 PASS`와 두 결과 파일을 확인합니다.",
        ),
        (
            "내 원본 대조 기준 입력",
            "OCR 원문에서 반드시 확인할 키워드 세 개를 정합니다.",
            "빈칸 여부 또는 내가 입력한 키워드가 표시되어야 합니다.",
        ),
        (
            "OCR 결과 묶음 완성",
            "공개 정답과 비교하고 OCR 산출물을 ZIP으로 묶습니다.",
            "대조 결과와 `lesson02_ocr_outputs.zip` 경로를 확인합니다.",
        ),
    ],
    "03_document_structure.ipynb": [
        (
            "공통 환경 준비",
            "이전 교시 산출물을 받을 폴더와 복구 기능을 준비합니다.",
            "Python·Platform·공통 작업 폴더가 표시되어야 합니다.",
        ),
        (
            "준비 입력 등록",
            "이전 산출물이 없을 때 사용할 공개 OCR 정답을 준비합니다.",
            "오류 없이 끝나면 준비 입력이 메모리에 등록된 것입니다.",
        ),
        (
            "공간 순서 복원 함수 준비",
            "OCR 좌표를 행과 읽기 순서로 재구성하는 함수를 만듭니다.",
            "오류 없이 끝나면 재구성 함수를 사용할 수 있습니다.",
        ),
        (
            "OCR를 문서 구조로 변환",
            "이전 OCR 결과를 읽고 헤더·품목·합계 후보로 나눕니다.",
            "입력 모드·원문 줄·품목 후보 수와 JSON 경로를 확인합니다.",
        ),
        (
            "내 품목 행 규칙 입력",
            "품목 행 끝의 단가·수량·금액 패턴을 정규식으로 표현합니다.",
            "내 규칙으로 찾은 품목 후보 수가 표시되어야 합니다.",
        ),
        (
            "품목 행 규칙 정답 확인",
            "공개 정규식으로 다섯 품목 행을 다시 찾습니다.",
            "다섯 품목 문자열이 목록으로 표시되어야 합니다.",
        ),
    ],
    "04_genai_extraction.ipynb": [
        (
            "공통 환경 준비",
            "이전 구조화 결과를 받을 폴더와 복구 기능을 준비합니다.",
            "Python·Platform·공통 작업 폴더가 표시되어야 합니다.",
        ),
        (
            "추출 정답과 비교 자료 준비",
            "영수증 원문·구조 초안·정답 스키마를 등록합니다.",
            "오류 없이 끝나면 비교 자료가 준비된 것입니다.",
        ),
        (
            "규칙 추출 함수 준비",
            "날짜·합계·품목·근거를 JSON으로 만드는 함수를 등록합니다.",
            "오류 없이 끝나면 추출 함수를 사용할 수 있습니다.",
        ),
        (
            "OCR 경로와 VLM 경로 비교",
            "이전 구조 결과를 읽어 두 경로의 결과와 출처를 구분합니다.",
            "총액·원문 근거·실행 모드와 두 JSON 파일을 확인합니다.",
        ),
        (
            "내 검토 결정 입력",
            "중요 필드·근거 유무·저장 전 행동을 직접 정합니다.",
            "빈칸 안내 또는 내가 내린 검토 결정이 표시되어야 합니다.",
        ),
        (
            "검토 정답 확인",
            "총액과 원문 근거를 기준으로 공개 검토 결정을 확인합니다.",
            "`REVIEW_BEFORE_EXPORT`가 포함된 정답을 확인합니다.",
        ),
    ],
    "05_streamlit_basic.ipynb": [
        (
            "공통 환경 준비",
            "웹앱 파일을 저장할 폴더와 실습 공통 기능을 준비합니다.",
            "Python·Platform·공통 작업 폴더가 표시되어야 합니다.",
        ),
        (
            "웹앱·OCR 실행환경 준비",
            "Streamlit과 PaddleOCR 버전을 확인하고 필요하면 설치합니다.",
            "웹앱 버전과 직접 OCR 실행 준비 결과가 표시되어야 합니다.",
        ),
        (
            "기본 웹앱 파일 생성",
            "업로드·실행 버튼·원문·JSON 영역이 있는 앱을 저장합니다.",
            "`app_05.py` 저장 경로가 표시되어야 합니다.",
        ),
        (
            "내 화면 문구 입력",
            "앱 제목과 실행 버튼을 업무 사용자가 이해할 말로 바꿉니다.",
            "빈칸 안내 또는 입력한 두 문구가 표시되어야 합니다.",
        ),
        (
            "화면 문구 적용",
            "내 문구 또는 공개 정답을 실제 앱 파일에 반영합니다.",
            "적용된 앱 제목과 버튼 문구를 확인합니다.",
        ),
        (
            "웹앱 자동 동작 검사",
            "업로드·버튼·준비 결과 화면이 실제로 동작하는지 검사합니다.",
            "`CHECKPOINT 1/1 PASS`가 표시되어야 합니다.",
        ),
        (
            "웹앱 직접 조작",
            "Colab 안에서 Streamlit 화면을 열어 버튼과 입력을 조작합니다.",
            "앱 화면 또는 검증 모드 생략 안내를 확인합니다.",
        ),
    ],
    "06_ocr_ai_integration.ipynb": [
        (
            "공통 환경 준비",
            "연동 앱 파일과 실습 자료를 위한 공통 환경을 준비합니다.",
            "Python·Platform·공통 작업 폴더가 표시되어야 합니다.",
        ),
        (
            "Streamlit 준비",
            "실습에 고정한 Streamlit 버전을 확인하고 필요하면 설치합니다.",
            "오류 없이 끝나면 웹앱 실행 환경이 준비된 것입니다.",
        ),
        (
            "OCR 연동 앱 생성",
            "직접 OCR·실패·수업용 예제 경로를 가진 앱 파일을 저장합니다.",
            "`app_06.py` 저장 경로가 표시되어야 합니다.",
        ),
        (
            "실제 OCR 기록 회귀검사",
            "보존한 PP-OCRv5 좌표를 행으로 복원하고 추출값을 검사합니다.",
            "`실제 OCR 기록 재검사 통과: 76000 5`를 확인합니다.",
        ),
        (
            "내 직접 실행 통과 조건 입력",
            "지금 OCR로 읽은 결과에서 반드시 맞아야 할 필드 세 개를 정합니다.",
            "빈칸 안내 또는 내가 입력한 세 필드가 표시되어야 합니다.",
        ),
        (
            "직접 실행 통과 조건 정답 확인",
            "날짜·총액·품목을 필수 확인값으로 확정합니다.",
            "세 필드가 담긴 전체 정답을 확인합니다.",
        ),
        (
            "연동 앱 자동 동작 검사",
            "준비 결과 버튼이 모드와 JSON을 화면에 표시하는지 검사합니다.",
            "`CHECKPOINT 1/1 PASS`가 표시되어야 합니다.",
        ),
        (
            "연동 앱 직접 조작",
            "Colab 안에서 직접 읽기와 수업용 예제 버튼을 조작합니다.",
            "앱 화면 또는 검증 모드 생략 안내를 확인합니다.",
        ),
    ],
    "07_validation_export.ipynb": [
        (
            "공통 환경 준비",
            "검증 결과와 Excel을 저장할 공통 환경을 준비합니다.",
            "Python·Platform·공통 작업 폴더가 표시되어야 합니다.",
        ),
        (
            "Streamlit 준비",
            "최종 앱에 필요한 Streamlit 버전을 확인하고 준비합니다.",
            "오류 없이 끝나면 최종 앱 실행 환경이 준비된 것입니다.",
        ),
        (
            "Excel 라이브러리 준비",
            "Excel 생성과 재검사에 필요한 openpyxl을 준비합니다.",
            "오류 없이 끝나면 Excel 기능을 사용할 수 있습니다.",
        ),
        (
            "검증용 정답 데이터 준비",
            "영수증 원문·품목·원문 근거가 포함된 정답을 등록합니다.",
            "오류 없이 끝나면 검증용 데이터가 준비된 것입니다.",
        ),
        (
            "이전 교시 결과 불러오기",
            "이전 JSON을 읽거나 공개 복구 입력으로 전환합니다.",
            "입력 모드와 검증 결과가 표시되어야 합니다.",
        ),
        (
            "검증·Excel 함수 준비",
            "오류·경고·승인 상태를 검사하고 Excel을 만드는 함수를 등록합니다.",
            "오류 없이 끝나면 검증과 저장 함수를 사용할 수 있습니다.",
        ),
        (
            "미승인 저장 차단 확인",
            "검토하지 않은 결과로 Excel이 생성되지 않는지 검사합니다.",
            "`DEFAULT_BLOCKED PASS`와 파일 미생성을 확인합니다.",
        ),
        (
            "내 승인 기록 입력",
            "승인 결정·검토자 이름·원본 확인 내용을 직접 입력합니다.",
            "빈칸 안내 또는 내가 남긴 승인 기록이 표시되어야 합니다.",
        ),
        (
            "승인 후 Excel 생성",
            "공개 승인 경로로 재검증하고 Excel 세 시트를 생성합니다.",
            "`REVIEWED_APPROVED PASS`와 Excel 파일 경로를 확인합니다.",
        ),
        (
            "최종 앱 묶음 생성",
            "업로드부터 Excel 다운로드까지 연결된 앱을 ZIP으로 묶습니다.",
            "최종 앱 경로와 ZIP 파일 경로가 표시되어야 합니다.",
        ),
        (
            "최종 앱 자동 동작 검사",
            "수정·재검증·승인 전 차단·승인 후 다운로드를 검사합니다.",
            "`FINAL APP PASS`가 표시되어야 합니다.",
        ),
        (
            "최종 앱 직접 조작",
            "Colab 안에서 전체 Document AI 흐름을 직접 수행합니다.",
            "앱 화면 또는 검증 모드 생략 안내를 확인합니다.",
        ),
    ],
    "08_business_application.ipynb": [
        (
            "공통 환경 준비",
            "업무 문서 샘플과 PoC 카드를 위한 공통 환경을 준비합니다.",
            "Python·Platform·공통 작업 폴더가 표시되어야 합니다.",
        ),
        (
            "실물형 문서 사진 비교",
            "견적서·신청서·거래명세서 사진을 차례로 관찰합니다.",
            "세 문서 이름·이미지 크기·미리보기가 보여야 합니다.",
        ),
        (
            "Office 형식과 확장 예제 체험",
            "Office 내부 구조를 검사하고 문서별 Python·JSON 예제를 묶습니다.",
            "Office ZIP과 `business_document_code_examples.zip`을 확인합니다.",
        ),
        (
            "내 PoC 후보 입력",
            "대상 문서·점수·검토자·중단 조건을 직접 정합니다.",
            "빈칸 안내 또는 내가 입력한 PoC 조건이 표시되어야 합니다.",
        ),
        (
            "PoC 후보 카드 완성",
            "점수 규칙으로 작은 PoC의 시작 여부를 판단해 저장합니다.",
            "후보 카드·`GO_SMALL` 또는 `REVIEW`·파일 경로를 확인합니다.",
        ),
    ],
}

LESSON_CODE_EXPLANATIONS = {
    "01_document_ai_overview.ipynb": [
        (
            "`OUTPUT_DIR`는 결과 파일을 모으는 폴더이고 `VALIDATION_MODE`는 "
            "자동검증에서만 사용합니다. 이 셀은 수정하지 않고 실행합니다."
        ),
        (
            "`load_course_assets()`는 먼저 GitHub에서 자료를 받고, 실패하면 "
            "Colab 업로드 창을 엽니다. 함수 내부는 수정하지 않습니다."
        ),
        (
            "`RECEIPT_IMAGE_PATH`는 이미지, `OCR_RECORD_PATH`는 OCR 좌표, "
            "`OCR_METADATA_PATH`는 좌표를 만든 이미지 크기와 해시입니다. "
            "세 값이 같은 원본에서 왔는지 확인한 뒤 표시합니다."
        ),
        (
            "`MY_SOURCE_OBSERVATION`이 내 답안입니다. 네 개의 `None`만 원본에서 "
            "찾은 값으로 바꾸고 나머지 코드는 그대로 둡니다."
        ),
        (
            "`ANSWER_SOURCE_OBSERVATION`은 공개 정답입니다. 리스트 컴프리헨션이 "
            "내 답과 정답을 한 행씩 비교하므로 수정하지 않고 실행합니다."
        ),
        (
            "`RECORDED_PP_OCRV5_TOKENS`에는 글자·좌표·신뢰도가 있습니다. "
            "`OCR_COORDINATE_SIZE`를 추측하지 않고 메타데이터에서 읽고, 원본과 "
            "가로세로 비율이 같은지 검사한 뒤 사각형을 그립니다."
        ),
        (
            "`MY_OCR_REVIEW`의 `None`만 `True` 또는 `False`로 바꿉니다. "
            "신뢰도와 업무 승인은 같은 뜻이 아니라는 점을 확인합니다."
        ),
        (
            "`VLM_DRAFT_WITH_ERROR`는 VLM 출력 모양을 흉내 낸 교육용 딕셔너리입니다. "
            "`items`의 합과 `total_amount`를 비교하며 읽습니다."
        ),
        (
            "`MY_VLM_REVIEW`의 `None`만 바꿉니다. JSON 구조가 자연스러운지와 "
            "그 안의 값이 정확한지를 서로 다른 질문으로 판단합니다."
        ),
        (
            "`validate_candidate()`는 오류 목록을 반환합니다. 품목 합계, 원본의 "
            "확인값, 원문 근거가 모두 맞아야 `valid=True`가 됩니다."
        ),
        (
            "`MY_CORRECTED_TOTAL`과 `MY_TOTAL_EVIDENCE` 두 곳만 입력합니다. "
            "수정한 복사본을 다시 검증하므로 원본 초안은 바뀌지 않습니다."
        ),
        (
            "`ANSWER_CORRECTED_TOTAL`과 `ANSWER_TOTAL_EVIDENCE`는 공개 정답입니다. "
            "수정 전·후 검증 결과가 달라지는 과정을 확인합니다."
        ),
        (
            "`MY_REVIEW_DECISION`에는 원본을 확인했으면 `APPROVED`를 입력합니다. "
            "자동검증 통과와 사람의 승인 기록이 별도 조건임을 보여 줍니다."
        ),
        (
            "`MY_CONCEPT_CHOICES`의 다섯 `None`만 기술 이름으로 바꿉니다. "
            "상황이 설명하는 역할을 보고 OCR·VLM·Document AI 등을 고릅니다."
        ),
        (
            "`comparison_report`는 앞 단계의 내 답과 정답 결과를 한 JSON에 모읍니다. "
            "`write_text()`가 `course_outputs`에 최종 파일을 저장합니다."
        ),
    ],
    "02_ocr_basic.ipynb": [
        (
            "`OUTPUT_DIR`는 모든 산출물의 공통 폴더입니다. 업로드·다운로드 함수와 "
            "자료 로더를 등록하는 준비 셀이므로 수정하지 않습니다."
        ),
        (
            "`USE_MY_FILE=False`면 공개 한국 영수증을 사용합니다. 내 사진이나 "
            "인터넷에서 찾은 공개 문서 이미지를 시험할 때만 `True`로 바꿉니다."
        ),
        (
            "`RUN_OCR_NOW`는 현재 이미지를 실제 OCR로 읽을지 정합니다. "
            "학생은 바꾸지 않고 실행하며 화면의 처리 결과만 확인합니다."
        ),
        (
            "`OCR_RESULT`를 원본 위에 그리고 인식 글자·신뢰도 표를 바로 표시합니다. "
            "`PROCESSING_PATH`를 함께 저장해 현재 실행과 이전 기록을 구분합니다."
        ),
        (
            "`review_keywords`의 세 `None`만 원본 대조에 사용할 문자열로 바꿉니다. "
            "OCR 원문에서 반드시 찾아야 할 값을 고르는 연습입니다."
        ),
        (
            "`ANSWER_REVIEW_KEYWORDS`는 공개 정답이고 `zipfile`은 JSON과 위치 이미지를 "
            "하나의 다운로드 파일로 묶습니다. 수정 없이 실행합니다."
        ),
    ],
    "03_document_structure.ipynb": [
        (
            "이전 교시의 `ocr_result.json`을 받을 공통 폴더와 업로드 함수를 준비합니다. "
            "설정 코드이므로 수정하지 않습니다."
        ),
        (
            "`GOLDEN_OCR_TEXT`와 `GOLDEN_RECEIPT`는 파일 인계가 막힐 때만 쓰는 "
            "공개 복구 데이터입니다. 수업용 예제와 지금 실행한 결과를 구분합니다."
        ),
        (
            "`reconstruct_spatial_lines()`는 y좌표가 가까운 토큰을 같은 행으로 묶고 "
            "x좌표 순서로 정렬합니다. OCR 글자를 읽기 순서로 복원하는 함수입니다."
        ),
        (
            "`USE_COURSE_EXAMPLE=True`면 새 Colab에서 공개 입력을 씁니다. 앞 교시 "
            "파일을 이어 쓰려면 `False`로 바꾸며, `groups`가 문서 영역을 나눕니다."
        ),
        (
            "`my_item_rule` 한 곳만 정규식으로 채웁니다. 숫자 세 묶음으로 끝나는 "
            "행을 품목 후보로 찾는 규칙입니다."
        ),
        (
            "`ANSWER_ITEM_RULE`은 공개 정규식입니다. `re.search()`가 각 행에서 "
            "규칙과 일치하는 품목 다섯 개를 찾는지 확인합니다."
        ),
    ],
    "04_genai_extraction.ipynb": [
        (
            "3교시의 `clean_receipt.json`을 받을 폴더와 복구 기능을 준비합니다. "
            "설정 코드이므로 수정하지 않습니다."
        ),
        (
            "`GOLDEN_VLM_MARKDOWN`은 현재 모델 호출 결과가 아니라 비교용 구조 초안입니다. "
            "실제 실행 결과와 준비 예제를 혼동하지 않습니다."
        ),
        (
            "`extract_receipt_from_text()`는 정규식으로 날짜·합계·품목을 찾고 "
            "각 값의 원문 근거까지 함께 반환합니다."
        ),
        (
            "`USE_COURSE_EXAMPLE=True`는 공개 입력, `False`는 3교시 파일 업로드입니다. "
            "`source_mode`와 `vlm_demo_mode`를 보고 두 결과의 출처를 구분합니다."
        ),
        (
            "`my_review`의 세 `None`만 채웁니다. 중요 필드, 원문 근거 유무, "
            "Excel 저장 전 행동을 직접 결정합니다."
        ),
        (
            "`ANSWER_REVIEW`는 총액의 원문 근거가 있더라도 저장 전에 사람이 "
            "검토해야 한다는 공개 정답입니다. 수정 없이 실행합니다."
        ),
    ],
    "05_streamlit_basic.ipynb": [
        (
            "웹앱 파일을 저장할 `OUTPUT_DIR`와 Colab 공통 함수를 준비합니다. "
            "설정 코드이므로 수정하지 않습니다."
        ),
        (
            "`required_streamlit`과 설치 버전을 비교합니다. 버전이 다를 때만 "
            "`pip install`을 실행하므로 이 셀은 그대로 실행합니다."
        ),
        (
            "`app_code`는 Streamlit 화면의 전체 소스이고 `write_text()`가 "
            "`app_05.py`로 저장합니다. 긴 문자열은 앱 파일을 만드는 재료입니다."
        ),
        (
            "`my_app_title`과 `my_button_label` 두 `None`만 원하는 문구로 바꿉니다. "
            "Python 문법보다 사용자에게 보이는 표현을 설계하는 단계입니다."
        ),
        (
            "`replace()`가 기본 제목과 버튼 문구를 내 문구로 바꾼 뒤 앱 파일을 "
            "다시 저장합니다. 빈칸이면 공개 정답을 사용합니다."
        ),
        (
            "`AppTest`는 브라우저를 열지 않고 제목·업로드·버튼·결과 영역을 검사합니다. "
            "예외가 없고 준비 결과가 보이면 통과입니다."
        ),
        (
            "`subprocess.Popen()`이 Streamlit 서버를 열고 Colab iframe에 표시합니다. "
            "선택 실습이며 화면이 안 열려도 앞의 AppTest 결과는 유지됩니다."
        ),
    ],
    "06_ocr_ai_integration.ipynb": [
        (
            "OCR 연동 앱과 결과 파일을 위한 공통 폴더·자료 로더를 준비합니다. "
            "설정 코드이므로 수정하지 않습니다."
        ),
        (
            "Streamlit과 PaddleOCR의 고정 버전을 확인하고 필요한 경우에만 설치합니다. "
            "설치 실패 시 준비 결과로 계속할 수 있도록 실패 이유를 보존합니다."
        ),
        (
            "`app_code`에는 업로드 파일을 OCR 함수에 전달하는 전체 앱 코드가 있습니다. "
            "`write_text()`가 이를 `app_06.py`로 저장합니다."
        ),
        (
            "`reconstruct_spatial_lines()`가 실제 OCR 좌표를 행으로 복원한 뒤 "
            "`extract_receipt_from_text()` 결과의 날짜·총액·품목 수를 검사합니다."
        ),
        (
            "`my_ocr_checks`의 세 `None`만 필드명으로 바꿉니다. 앱이 열리는 것과 "
            "추출값이 맞는 것은 다르므로 업무 통과 조건을 고릅니다."
        ),
        (
            "`ANSWER_OCR_CHECKS`는 날짜·총액·품목을 필수 확인값으로 제시합니다. "
            "내 답과 비교하고 수정 없이 실행합니다."
        ),
        (
            "`AppTest`가 수업용 예제 버튼을 누르고 처리 방식과 JSON이 화면에 "
            "나오는지 검사합니다. 실제 OCR 정확도 검사는 앞의 회귀 셀이 담당합니다."
        ),
        (
            "Streamlit 서버를 Colab iframe으로 열어 직접 읽기와 수업용 예제 버튼을 조작합니다. "
            "서버 프로세스와 자동검증 분기를 수정하지 않습니다."
        ),
    ],
    "07_validation_export.ipynb": [
        (
            "검증 결과와 Excel을 저장할 공통 폴더·파일 인계 함수를 준비합니다. "
            "설정 코드이므로 수정하지 않습니다."
        ),
        (
            "최종 앱에서 사용할 Streamlit 버전을 확인하고 필요할 때만 설치합니다. "
            "버전 숫자는 수업 중 임의로 바꾸지 않습니다."
        ),
        (
            "`find_spec()`로 openpyxl 설치 여부를 확인합니다. 없을 때만 설치해 "
            "Excel 생성 함수를 사용할 수 있게 합니다."
        ),
        (
            "`GOLDEN_RECEIPT`는 검증 규칙과 Excel 구조를 확인할 공개 정답입니다. "
            "이 셀은 데이터를 등록하므로 수정하지 않습니다."
        ),
        (
            "`USE_COURSE_EXAMPLE=True`는 공개 입력, `False`는 4교시 JSON 업로드입니다. "
            "`validate_receipt()` 결과의 오류와 경고를 확인합니다."
        ),
        (
            "`validate_receipt()`는 필수값·계산·근거를 검사하고 "
            "`save_reviewed_excel()`은 "
            "검증과 승인 조건이 모두 맞을 때만 세 시트를 만듭니다."
        ),
        (
            "`PENDING_REVIEW` 상태로 `save_reviewed_excel()`을 호출합니다. "
            "반환값이 `False`이고 "
            "파일이 없으면 미승인 저장 차단이 정상입니다."
        ),
        (
            "`my_decision`, `my_reviewer`, `my_review_note` 세 곳만 입력합니다. "
            "원본 대조를 끝낸 뒤 실제 검토 기록을 남기는 단계입니다."
        ),
        (
            "`reviewed_receipt`에 공개 샘플의 원본 대조 정답을 적용하고 "
            "`REVIEW_RECORD`와 함께 Excel을 만듭니다. 내 자료는 내가 승인 기록을 "
            "채우기 전까지 저장하지 않습니다."
        ),
        (
            "`FINAL_APP_SOURCE_PATHS`는 최종 앱에 필요한 파일 목록입니다. "
            "`load_course_assets()`로 파일을 받아 폴더에 저장한 뒤 "
            "`make_archive()`가 ZIP으로 묶습니다."
        ),
        (
            "`AppTest`가 오류값 저장 차단, 정상값 복구, 사람 승인, Excel 다운로드를 "
            "버튼 조작으로 검사합니다. 최종 통합 테스트입니다."
        ),
        (
            "최종 Streamlit 앱을 Colab iframe으로 열어 전체 흐름을 직접 조작합니다. "
            "자동검증에서는 서버 화면만 생략합니다."
        ),
    ],
    "08_business_application.ipynb": [
        (
            "업무 문서 샘플과 PoC 결과를 저장할 공통 폴더·자료 로더를 준비합니다. "
            "설정 코드이므로 수정하지 않습니다."
        ),
        (
            "`EXTENSION_IMAGE_PATHS`가 문서별 사진 경로를 연결합니다. 반복문은 "
            "세 이미지를 같은 크기로 줄여 비교하기 쉽게 표시합니다."
        ),
        (
            "`xml_text_count()`는 Office 내부 XML을 셉니다. 이어서 견적서·신청서·"
            "거래명세서의 검증 Python과 정답 JSON을 별도 ZIP으로 묶습니다."
        ),
        (
            "`candidate`, `score`, `review_owner`, `stop_condition`의 빈칸만 채웁니다. "
            "모델 정확도보다 작은 PoC를 운영할 조건을 정하는 단계입니다."
        ),
        (
            "점수 조건이 맞으면 `GO_SMALL`, 아니면 `REVIEW`를 선택합니다. "
            "결과와 검토·중단 조건을 Markdown PoC 카드로 저장합니다."
        ),
    ],
}

FINAL_APP_SOURCE_PATHS = [
    "app.py",
    "src/__init__.py",
    "src/clean.py",
    "src/export.py",
    "src/extract.py",
    "src/ocr.py",
    "src/pipeline.py",
    "src/sample_data.py",
    "src/validate.py",
    "src/vlm.py",
]


def markdown(source: str) -> dict:
    return {
        "cell_type": "markdown",
        "id": "pending",
        "metadata": {},
        "source": dedent(source).strip() + "\n",
    }


def code(source: str) -> dict:
    return {
        "cell_type": "code",
        "execution_count": None,
        "id": "pending",
        "metadata": {},
        "outputs": [],
        "source": dedent(source).strip() + "\n",
    }


def learning_progress_source() -> str:
    """각 코드 셀의 시작·완료 상태를 Colab 출력으로 분명하게 보여 준다."""

    return dedent(
        '''
        def _show_learning_message(markdown_text):
            try:
                from IPython.display import Markdown, display
                display(Markdown(markdown_text))
            except ImportError:
                print(markdown_text)


        def show_lab_step(
            current,
            total,
            title,
            action,
            expected,
            code_help,
            edit_kind,
        ):
            cell_kind = {
                "required": "🟠 내가 짧게 바꾸는 셀",
                "optional": "🔵 원하면 바꾸는 셀",
                "none": "🟢 그대로 실행하는 셀",
            }[edit_kind]
            _show_learning_message(
                f"""---
        ### {cell_kind} · {current}/{total} · {title}

        **지금 할 일:** {action}

        **코드 읽는 법:** {code_help}

        **이 단계에서 확인할 결과:** {expected}
        """
            )


        def complete_lab_step(current, total, expected):
            next_action = (
                "결과를 확인한 뒤 다음 코드 셀을 실행하세요."
                if current < total
                else "마지막 CHECKPOINT와 산출물 파일을 확인하세요."
            )
            _show_learning_message(
                f"""> ✅ **{current}/{total} 단계 실행 완료**
        >
        > **결과 확인:** {expected}
        >
        > **다음 행동:** {next_action}
        """
            )
        '''
    ).strip()


def attach_learning_guides(name: str, cells: list[dict]) -> None:
    guides = LESSON_STEP_GUIDES[name]
    explanations = LESSON_CODE_EXPLANATIONS[name]
    code_cells = [cell for cell in cells if cell["cell_type"] == "code"]
    if len(code_cells) != len(guides):
        raise ValueError(
            f"{name}: 코드 셀 {len(code_cells)}개와 단계 안내 "
            f"{len(guides)}개가 다릅니다."
        )
    if len(code_cells) != len(explanations):
        raise ValueError(
            f"{name}: 코드 셀 {len(code_cells)}개와 코드 설명 "
            f"{len(explanations)}개가 다릅니다."
        )

    total = len(code_cells)
    for current, (cell, guide, explanation) in enumerate(
        zip(code_cells, guides, explanations),
        start=1,
    ):
        title, action, expected = guide
        cell["metadata"]["learning_step"] = {
            "current": current,
            "total": total,
            "title": title,
            "action": action,
            "expected": expected,
            "code_help": explanation,
            "learner_edits": "# TODO" in cell["source"],
            "edit_kind": (
                "optional"
                if "# TODO(선택)" in cell["source"]
                else "required"
                if "# TODO" in cell["source"]
                else "none"
            ),
        }
        comment_lines = [
            "# ── 코드 읽기 ─────────────────────────────────────────────",
        ]
        for line in wrap(explanation, width=82):
            comment_lines.append(f"# {line}")
        comment_lines.append(
            "# ──────────────────────────────────────────────────────────"
        )
        comment_block = "\n".join(comment_lines)
        edit_kind = (
            "optional"
            if "# TODO(선택)" in cell["source"]
            else "required"
            if "# TODO" in cell["source"]
            else "none"
        )
        start = (
            f"show_lab_step({current}, {total}, {title!r}, "
            f"{action!r}, {expected!r}, {explanation!r}, {edit_kind!r})"
        )
        finish = f"complete_lab_step({current}, {total}, {expected!r})"
        original = cell["source"].rstrip()
        if current == 1:
            original = (
                learning_progress_source()
                + "\n\n"
                + comment_block
                + "\n"
                + start
                + "\n\n"
                + original
            )
        else:
            original = comment_block + "\n" + start + "\n\n" + original
        cell["source"] = original + "\n\n" + finish + "\n"


def notebook(name: str, cells: list[dict]) -> dict:
    first_markdown = next(
        cell for cell in cells if cell["cell_type"] == "markdown"
    )
    beginner_path = LESSON_BEGINNER_PATHS[name]
    first_markdown["source"] += dedent(
        f"""

        ## 이 노트북에서 내가 하는 일

        - **필수 실습:** {beginner_path["required"]}
        - **내가 바꾸는 곳:** {beginner_path["edit"]}
        - **인터넷 자료로 다시 실험:** {beginner_path["research"]}

        먼저 제공 샘플로 끝까지 실행해 `CHECKPOINT PASS`를 만드세요. 그다음
        [공개·비식별 실습 자료 찾기](https://github.com/leecks1119/document_ai_lecture/blob/master/docs/public_practice_sources.md)를 보고
        입력 한 장만 바꾸어 다시 실행합니다. 2교시에서 고른 자료와 결과 파일은
        3~7교시에 그대로 이어 쓰므로 매 시간 새 자료를 찾을 필요가 없습니다.

        > `🟢 그대로 실행하는 셀`은 수정하지 않습니다. `🟠 내가 짧게 바꾸는
        > 셀`만 필수이고, `🔵 원하면 바꾸는 셀`은 시간이 남을 때 합니다.
        > 정답은 모두 공개되어 있으므로 정답을 먼저 복사하고 결과를 관찰해도 됩니다.

        ## 코드 셀을 읽는 방법

        각 코드 셀의 맨 위에는 `코드 읽기` 주석이 있습니다.

        1. `수정하지 않습니다`라고 적힌 셀은 설명을 읽고 그대로 실행합니다.
        2. 주황색 필수 `TODO`만 채웁니다. 파란색 선택 `TODO`는 건너뛰어도 됩니다.
        3. 실행 출력에서 `코드 읽는 법`과 `확인할 결과`를 다시 확인합니다.
        4. `단계 실행 완료`가 나온 뒤 다음 코드 셀로 이동합니다.

        Python 문법 전체를 먼저 이해할 필요는 없습니다. 변수에 어떤 값이 들어가고,
        실행 뒤 어떤 결과가 달라지는지를 중심으로 읽습니다.
        """
    )
    attach_learning_guides(name, cells)
    prefix = name[:2]
    for index, cell in enumerate(cells, start=1):
        cell["id"] = f"{prefix}-{index:02d}"
    return {
        "cells": cells,
        "metadata": {
            "colab": {"name": name, "provenance": []},
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {"name": "python"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }


def intro(lesson: int, title: str, artifact: str, goal: str) -> dict:
    slug = NOTEBOOK_SLUGS[lesson]
    return markdown(
        f"""
        # {lesson}교시. {title}

        [![Open In Colab](https://colab.research.google.com/assets/colab-badge.svg)](https://colab.research.google.com/github/leecks1119/document_ai_lecture/blob/master/colab/{lesson:02d}_{slug}.ipynb)

        **이번 교시 행동:** {goal}

        **통과 증거:** `course_outputs/{artifact}`

        > Google Colab도 외부 클라우드입니다. 조직 승인 없는 개인·회사 문서는
        > 업로드하지 않습니다. 필수 실습은 저장소의 비식별 공개·합성 샘플만
        > 사용합니다.

        화면의 **처리 방식**을 먼저 확인합니다.

        - **지금 이 사진을 직접 읽었습니다:** 현재 파일에 OCR 모델을 실행한 결과입니다.
        - **수업용 예제 결과를 불러왔습니다:** 현재 파일을 분석한 결과가 아닙니다.
        - 3분 이상 멈추면 실행을 중지하고 수업용 예제로 계속합니다.
        - 각 교시 끝에서 `CHECKPOINT PASS`와 산출물 파일을 확인합니다.
        """
    )


def course_asset_loader_source() -> str:
    """Colab에서는 URL, 자동 검증에서는 저장소 파일로 실습 자료를 읽는다."""

    return dedent(
        """
        COURSE_ASSET_BASE_URL = (
            "https://raw.githubusercontent.com/leecks1119/"
            "document_ai_lecture/master/"
        )

        def load_course_assets(*relative_paths):
            if VALIDATION_MODE:
                local_root = os.getenv("COURSE_LOCAL_ASSET_ROOT")
                if not local_root:
                    raise RuntimeError(
                        "자동 검증용 COURSE_LOCAL_ASSET_ROOT가 필요합니다."
                    )
                root = Path(local_root)
                return {
                    path: (root / path).read_bytes()
                    for path in relative_paths
                }

            import requests

            loaded = {}
            missing = []
            for path in relative_paths:
                try:
                    response = requests.get(
                        COURSE_ASSET_BASE_URL + path,
                        timeout=30,
                    )
                    response.raise_for_status()
                    loaded[path] = response.content
                except requests.RequestException as exc:
                    print(f"자동 다운로드 실패: {Path(path).name} · {exc}")
                    missing.append(path)

            if missing:
                from google.colab import files

                expected = ", ".join(Path(path).name for path in missing)
                print("다음 파일을 저장소에서 내려받아 선택하세요:", expected)
                uploaded = files.upload()
                uploaded_by_name = {
                    Path(name).name: content
                    for name, content in uploaded.items()
                }
                for path in missing:
                    filename = Path(path).name
                    if filename not in uploaded_by_name:
                        raise FileNotFoundError(
                            f"{filename}이 선택되지 않았습니다."
                        )
                    loaded[path] = uploaded_by_name[filename]

            return loaded
        """
    ).strip()


def runtime_cell() -> dict:
    setup = dedent(
        """
        import json
        import os
        import platform
        import sys
        from pathlib import Path

        OUTPUT_DIR = Path("course_outputs")
        OUTPUT_DIR.mkdir(exist_ok=True)
        VALIDATION_MODE = os.getenv("COURSE_VALIDATE_EXAMPLE") == "1"

        def upload_previous_artifact(filename):
            target = OUTPUT_DIR / filename
            if target.exists() or VALIDATION_MODE:
                return target if target.exists() else None
            try:
                from google.colab import files
            except ImportError:
                return None
            print(f"이전 교시에서 내려받은 {filename}을 선택하세요.")
            uploaded = files.upload()
            if filename not in uploaded:
                raise FileNotFoundError(
                    f"{filename}이 선택되지 않았습니다. 준비 입력을 쓰려면 "
                    "USE_COURSE_EXAMPLE=True로 바꾸세요."
                )
            target.write_bytes(uploaded[filename])
            return target


        def download_artifact(path):
            if VALIDATION_MODE:
                return
            try:
                from google.colab import files
            except ImportError:
                return
            files.download(str(path))

        print("Python:", sys.version.split()[0])
        print("Platform:", platform.platform())
        print("공통 작업 폴더:", OUTPUT_DIR.resolve())
        """
    ).strip()
    return code(setup + "\n\n" + course_asset_loader_source())


def streamlit_preview_cell(path_expression: str, port: int) -> dict:
    return code(
        f"""
        # 선택 실습 · 녹화에서는 이 셀로 실제 화면을 엽니다.
        # AppTest가 필수 검증이며, 미리보기에는 공개 비식별 샘플만 사용합니다.
        if not VALIDATION_MODE:
            import subprocess
            import time
            import urllib.request

            preview_process = subprocess.Popen(
                [
                    sys.executable, "-m", "streamlit", "run",
                    str({path_expression}),
                    "--server.port", "{port}",
                    "--server.headless", "true",
                    "--server.enableCORS", "false",
                    "--server.enableXsrfProtection", "false",
                ],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.STDOUT,
            )
            for _ in range(20):
                try:
                    urllib.request.urlopen(
                        "http://127.0.0.1:{port}/_stcore/health",
                        timeout=1,
                    )
                    break
                except Exception:
                    time.sleep(0.5)
            try:
                from google.colab import output
                print("아래 화면에서 직접 버튼과 입력값을 조작하세요.")
                output.serve_kernel_port_as_iframe({port}, height=760)
            except Exception as exc:
                print("Colab 미리보기를 열지 못했습니다:", exc)
                print("AppTest 결과와 app 파일로 계속합니다.")
        else:
            print("검증 모드: 대화형 Streamlit 미리보기 생략")
        """
    )


def readable_string_assignment(name: str, value: str) -> str:
    """긴 문자열을 노트북에서 읽을 수 있는 인접 문자열 형태로 만든다."""

    lines = value.splitlines(keepends=True)
    if not lines:
        return f"{name} = ''"
    body = "\n".join(f"    {line!r}" for line in lines)
    return f"{name} = (\n{body}\n)"

def golden_constants() -> str:
    return (
        f"GOLDEN_OCR_TEXT = {GOLDEN_OCR_TEXT!r}\n"
        f"GOLDEN_VLM_MARKDOWN = {GOLDEN_VLM_MARKDOWN!r}\n"
        f"GOLDEN_RECEIPT = {pformat(GOLDEN_RECEIPT, sort_dicts=False, width=88)}\n"
    )


def parser_source() -> str:
    return r'''
import re

def to_int(value):
    return int(value.replace(",", ""))


def extract_receipt_from_text(text, source_mode):
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    date_match = re.search(r"\b(\d{4})[-./](\d{1,2})[-./](\d{1,2})\b", text)
    total_line = next(
        (
            line
            for line in lines
            if re.search(r"(?:합\s*계|결제\s*금액|총\s*액)", line)
        ),
        None,
    )
    total_candidates = (
        re.findall(r"(?<![\d,])\d[\d,]*(?![\d,])", total_line)
        if total_line
        else []
    )
    total_raw = total_candidates[-1] if total_candidates else None
    supply_match = re.search(
        r"(?:부가세\s*)?과세물품가액\s*[:：]?\s*([\d,]+)",
        text,
    )
    vat_match = re.search(
        r"^부가세(?!\s*과세물품가액)\s*[:：]?\s*([\d,]+)",
        text,
        re.MULTILINE,
    )
    item_pattern = re.compile(
        r"^(?P<name>.+?)\s+(?P<unit>[\d,]+)\s+"
        r"(?P<quantity>\d+)\s+(?P<line>[\d,]+)$"
    )
    markdown_item_pattern = re.compile(
        r"^\|\s*(?P<name>[^|]+?)\s*\|\s*(?P<quantity>\d+)\s*\|"
        r"\s*(?P<unit>[\d,]+)원\s*\|\s*(?P<line>[\d,]+)원\s*\|$"
    )
    items = []
    item_evidence = []
    for line_number, line in enumerate(lines, start=1):
        match = item_pattern.search(line)
        if not match:
            match = markdown_item_pattern.search(line)
        if match:
            item = {
                "name": match.group("name"),
                "quantity": int(match.group("quantity")),
                "unit_price": to_int(match.group("unit")),
                "line_total": to_int(match.group("line")),
            }
            items.append(item)
            item_evidence.append({"line": line_number, "raw_value": line})

    date_value = (
        f"{int(date_match.group(1)):04d}-{int(date_match.group(2)):02d}-"
        f"{int(date_match.group(3)):02d}"
        if date_match else None
    )
    total_value = to_int(total_raw) if total_raw else None
    supply_value = to_int(supply_match.group(1)) if supply_match else None
    vat_value = to_int(vat_match.group(1)) if vat_match else None
    return {
        "document_type": "receipt",
        "store_name": lines[0] if lines else None,
        "date": date_value,
        "total_amount": total_value,
        "items": items,
        "adjustments": {"discount": 0, "tax": 0, "service": 0, "rounding": 0},
        "tax_breakdown": {
            "mode": "included_in_item_prices",
            "supply_amount": supply_value,
            "vat": vat_value,
            "payable_total": total_value,
        } if supply_value is not None and vat_value is not None else None,
        "raw_values": {
            "store_name": lines[0] if lines else None,
            "date": date_match.group(0) if date_match else None,
            "total_amount": total_raw,
        },
        "cleaned_values": {
            "store_name": lines[0] if lines else None,
            "date": date_value,
            "total_amount": total_value,
        },
        "evidence": {
            "store_name": {"line": 1, "raw_value": lines[0] if lines else None},
            "date": {"raw_value": date_match.group(0) if date_match else None},
            "total_amount": {"raw_value": total_line},
            "items": item_evidence,
        },
        "source_mode": source_mode,
    }
'''


def spatial_reconstruction_source() -> str:
    return r'''
from collections import defaultdict

def reconstruct_spatial_lines(items):
    positioned_by_page = defaultdict(list)
    unpositioned_by_page = defaultdict(list)
    for order, item in enumerate(items):
        text = " ".join(str(item.get("text", "")).split())
        if not text:
            continue
        page = int(item.get("page") or 1)
        points = [
            point
            for point in (item.get("box") or [])
            if isinstance(point, (list, tuple)) and len(point) >= 2
        ]
        if not points:
            unpositioned_by_page[page].append((order, text))
            continue
        xs = [float(point[0]) for point in points]
        ys = [float(point[1]) for point in points]
        positioned_by_page[page].append({
            "text": text,
            "x": min(xs),
            "y": sum(ys) / len(ys),
            "height": max(ys) - min(ys),
            "order": order,
        })

    pages = sorted(set(positioned_by_page) | set(unpositioned_by_page))
    lines = []
    for page in pages:
        rows = []
        for token in sorted(
            positioned_by_page[page],
            key=lambda value: (value["y"], value["x"], value["order"]),
        ):
            row = rows[-1] if rows else None
            tolerance = (
                max(12.0, min(24.0, max(row["height"], token["height"]) * 0.45))
                if row else 12.0
            )
            if row and abs(token["y"] - row["y"]) <= tolerance:
                row["tokens"].append(token)
                count = len(row["tokens"])
                row["y"] = (row["y"] * (count - 1) + token["y"]) / count
                row["height"] = max(row["height"], token["height"])
            else:
                rows.append({
                    "tokens": [token],
                    "y": token["y"],
                    "height": token["height"],
                })

        lines.extend(
            " ".join(
                token["text"]
                for token in sorted(
                    row["tokens"],
                    key=lambda value: (value["x"], value["order"]),
                )
            )
            for row in rows
        )
        lines.extend(
            text
            for _, text in sorted(
                unpositioned_by_page[page],
                key=lambda value: value[0],
            )
        )
    return lines
'''


def notebook_01() -> dict:
    course_example_items = repr(GOLDEN_RECEIPT["items"])
    source_image_cell = code(
        """
        RECEIPT_IMAGE_PATH = (
            "sample_docs/public_receipts/korea/"
            "taebaek_restaurant_2025_redacted.png"
        )
        OCR_RECORD_PATH = "tests/fixtures/ppocrv5_recorded_receipt_tokens.json"
        OCR_METADATA_PATH = "tests/fixtures/ppocrv5_recorded_receipt_metadata.json"
        lesson_assets = load_course_assets(
            RECEIPT_IMAGE_PATH,
            OCR_RECORD_PATH,
            OCR_METADATA_PATH,
        )
        receipt_image = Image.open(
            io.BytesIO(lesson_assets[RECEIPT_IMAGE_PATH])
        ).convert("RGB")
        RECORDED_PP_OCRV5_TOKENS = json.loads(
            lesson_assets[OCR_RECORD_PATH].decode("utf-8")
        )
        OCR_RECORD_METADATA = json.loads(
            lesson_assets[OCR_METADATA_PATH].decode("utf-8")
        )
        assert hashlib.sha256(
            lesson_assets[RECEIPT_IMAGE_PATH]
        ).hexdigest() == OCR_RECORD_METADATA["source_image_sha256"]
        assert hashlib.sha256(
            lesson_assets[OCR_RECORD_PATH]
        ).hexdigest() == OCR_RECORD_METADATA["token_file_sha256"]
        assert receipt_image.size == (
            OCR_RECORD_METADATA["source_image_size"]["width"],
            OCR_RECORD_METADATA["source_image_size"]["height"],
        )
        image_source = (
            "자동 검증용 저장소 파일"
            if VALIDATION_MODE
            else "저장소에서 내려받은 공개 영수증"
        )

        preview = receipt_image.copy()
        preview.thumbnail((540, 660))
        print("이미지 출처:", image_source)
        display(preview)
        """
    )
    cells = [
        markdown(
            """
            # 1교시. 한국 영수증으로 구분하는 OCR·VLM·Document AI

            [![Open In Colab](https://colab.research.google.com/assets/colab-badge.svg)](https://colab.research.google.com/github/leecks1119/document_ai_lecture/blob/master/colab/01_document_ai_overview.ipynb)

            ## 오늘의 도전

            같은 영수증을 세 가지 결과로 살펴봅니다.

            1. 실제 PP-OCRv5 실행에서 기록한 **글자·위치·신뢰도**
            2. 이미지와 지시문을 함께 본다고 가정한 **VLM 구조 초안**
            3. 원본 근거와 계산 규칙으로 초안을 검사한 **Document AI 결과**

            일부러 잘못 넣은 합계 `16,000원`을 찾아 `76,000원`으로 고치고,
            사람 확인 전에는 Excel로 보낼 수 없는 이유까지 확인합니다.

            > Google Colab도 외부 클라우드입니다. 필수 실습은 개인정보를 가린
            > 공개 한국 영수증을 사용합니다. 조직 승인 없는 개인·회사 문서는
            > 업로드하지 않습니다.

            이번 시간에는 모델 설치에 시간을 쓰지 않습니다. OCR 화면은 이 영수증에
            PP-OCRv5를 실제 실행해 보존한 결과이고, VLM 화면은 개념 비교를 위해
            오류를 넣어 만든 교육용 구조 초안입니다. 실제 OCR 실행은 2교시에서 진행합니다.
            """
        ),
        code(
            """
            import copy
            import hashlib
            import io
            import json
            import os
            from pathlib import Path

            import pandas as pd
            from PIL import Image
            from PIL import ImageDraw

            try:
                from IPython.display import display
            except ImportError:
                def display(value):
                    print(value)

            OUTPUT_DIR = Path("course_outputs")
            OUTPUT_DIR.mkdir(exist_ok=True)
            VALIDATION_MODE = os.getenv("COURSE_VALIDATE_EXAMPLE") == "1"

            print("실습 준비 완료")
            """
        ),
        code(course_asset_loader_source()),
        markdown(
            """
            ## 1. 먼저 사람의 눈으로 원본을 읽습니다

            AI 결과를 보기 전에 원본에서 다음 네 가지를 직접 찾아보세요.

            - 상호명
            - 거래 날짜
            - 품목 수
            - 합계 금액
            """
        ),
        source_image_cell,
        code(
            """
            # TODO: 영수증 원본을 보고 None 네 곳을 채우세요.
            MY_SOURCE_OBSERVATION = {
                "store_name": None,
                "date": None,
                "item_count": None,
                "total_amount": None,
            }
            display(pd.DataFrame([MY_SOURCE_OBSERVATION]))
            """
        ),
        markdown(
            """
            <details>
            <summary>힌트 보기</summary>

            합계는 영수증 아래쪽의 `합계` 행에 있습니다. 품목은 표 안에서
            다섯 줄을 찾습니다. 금액은 쉼표를 제외한 정수로 적습니다.
            </details>
            """
        ),
        code(
            """
            # 전체 정답: 원본을 다시 보면서 자신의 관찰과 비교하세요.
            ANSWER_SOURCE_OBSERVATION = {
                "store_name": "이태리집",
                "date": "2025-10-04",
                "item_count": 5,
                "total_amount": 76000,
            }
            SOURCE_LABELS = {
                "store_name": "상호명",
                "date": "거래 날짜",
                "item_count": "품목 수",
                "total_amount": "합계 금액",
            }
            source_comparison = pd.DataFrame(
                [
                    {
                        "항목": SOURCE_LABELS[key],
                        "내 관찰": MY_SOURCE_OBSERVATION[key],
                        "원본 확인값": value,
                        "일치": (
                            "미작성"
                            if MY_SOURCE_OBSERVATION[key] is None
                            else MY_SOURCE_OBSERVATION[key] == value
                        ),
                    }
                    for key, value in ANSWER_SOURCE_OBSERVATION.items()
                ]
            )
            display(source_comparison)
            """
        ),
        markdown(
            """
            ## 2. OCR 결과에서 글자·위치·신뢰도를 관찰합니다

            아래 결과는 같은 영수증에 PP-OCRv5를 실제 실행해 기록한 44개 글자
            조각입니다. 초록색 사각형은 신뢰도 0.8 이상, 빨간색 사각형은
            0.8 미만입니다.

            OCR은 글자를 잘 읽은 곳과 틀린 곳을 동시에 보여 줍니다. 특히
            상호명과 부가세 과세물품가액을 원본과 비교해 보세요.
            """
        ),
        code(
            """
            assert len(RECORDED_PP_OCRV5_TOKENS) == 44

            annotated_receipt = receipt_image.copy()
            draw = ImageDraw.Draw(annotated_receipt)
            OCR_COORDINATE_SIZE = (
                OCR_RECORD_METADATA["coordinate_space"]["width"],
                OCR_RECORD_METADATA["coordinate_space"]["height"],
            )
            expected_coordinate_height = round(
                receipt_image.height
                * OCR_COORDINATE_SIZE[0]
                / receipt_image.width
            )
            assert OCR_COORDINATE_SIZE[1] == expected_coordinate_height
            scale_x = annotated_receipt.width / OCR_COORDINATE_SIZE[0]
            scale_y = annotated_receipt.height / OCR_COORDINATE_SIZE[1]
            for token in RECORDED_PP_OCRV5_TOKENS:
                points = [
                    (
                        round(point[0] * scale_x),
                        round(point[1] * scale_y),
                    )
                    for point in token["box"]
                ]
                confidence = token["confidence"] or 0
                color = "#18A558" if confidence >= 0.8 else "#E34A33"
                draw.line(points + [points[0]], fill=color, width=4)

            annotated_preview = annotated_receipt.copy()
            annotated_preview.thumbnail((540, 660))
            print(
                "OCR 좌표 기준:",
                OCR_COORDINATE_SIZE,
                "→ 원본 이미지:",
                receipt_image.size,
            )
            display(annotated_preview)

            OCR_FOCUS = [
                {
                    "항목": "상호명",
                    "원본": "이태리집",
                    "OCR 판독": RECORDED_PP_OCRV5_TOKENS[2]["text"],
                    "신뢰도": round(RECORDED_PP_OCRV5_TOKENS[2]["confidence"], 3),
                },
                {
                    "항목": "거래일시",
                    "원본": "2025-10-04 12:33:37",
                    "OCR 판독": RECORDED_PP_OCRV5_TOKENS[4]["text"],
                    "신뢰도": round(RECORDED_PP_OCRV5_TOKENS[4]["confidence"], 3),
                },
                {
                    "항목": "합계",
                    "원본": "76,000",
                    "OCR 판독": RECORDED_PP_OCRV5_TOKENS[34]["text"],
                    "신뢰도": round(RECORDED_PP_OCRV5_TOKENS[34]["confidence"], 3),
                },
                {
                    "항목": "부가세 과세물품가액",
                    "원본": "69,094",
                    "OCR 판독": RECORDED_PP_OCRV5_TOKENS[36]["text"],
                    "신뢰도": round(RECORDED_PP_OCRV5_TOKENS[36]["confidence"], 3),
                },
            ]
            display(pd.DataFrame(OCR_FOCUS))
            """
        ),
        code(
            """
            # TODO(선택): 위 표와 원본을 비교해 True 또는 False로 채우세요.
            MY_OCR_REVIEW = {
                "상호명이 원본과 같다": None,
                "거래일시가 원본과 같다": None,
                "합계가 원본과 같다": None,
                "과세물품가액이 원본과 같다": None,
                "높은 신뢰도는 업무 승인을 뜻한다": None,
            }

            ANSWER_OCR_REVIEW = {
                "상호명이 원본과 같다": False,
                "거래일시가 원본과 같다": True,
                "합계가 원본과 같다": True,
                "과세물품가액이 원본과 같다": False,
                "높은 신뢰도는 업무 승인을 뜻한다": False,
            }

            ocr_review_result = pd.DataFrame(
                [
                    {
                        "확인 질문": key,
                        "내 답": MY_OCR_REVIEW[key],
                        "전체 정답": expected,
                        "결과": (
                            "미작성"
                            if MY_OCR_REVIEW[key] is None
                            else "맞음"
                            if MY_OCR_REVIEW[key] == expected
                            else "다시 원본 확인"
                        ),
                    }
                    for key, expected in ANSWER_OCR_REVIEW.items()
                ]
            )
            display(ocr_review_result)
            """
        ),
        markdown(
            """
            **관찰 결과**

            OCR은 합계 `76,000`을 높은 신뢰도로 읽었지만 상호명은
            `이태리쉽!`, 과세물품가액은 `639,094`로 잘못 읽었습니다.
            신뢰도는 모델의 판독 확신이지 업무 승인 점수가 아닙니다.
            """
        ),
        markdown(
            """
            ## 3. VLM 구조 초안에서 문맥과 오류를 함께 봅니다

            VLM은 문서 이미지와 “상호명·날짜·품목·합계를 JSON으로 정리하라”는
            지시문을 함께 받을 수 있습니다. 아래 결과는 VLM이 일반적으로 만드는
            구조를 이해하기 위한 **교육용 예제**이며, 현재 Colab에서 모델을 호출한
            결과가 아닙니다.

            표 구조는 잘 정리되어 있지만 합계에 일부러 오류가 들어 있습니다.
            원본과 품목 금액을 이용해 오류를 찾아보세요.
            """
        ),
        code(
            f"""
            VLM_DRAFT_WITH_ERROR = {{
                "store_name": "이태리집",
                "date": "2025-10-04",
                "items": {course_example_items},
                "total_amount": 16000,
                "evidence": {{}},
            }}

            display(
                pd.DataFrame(
                    [
                        {{
                            "상호명": VLM_DRAFT_WITH_ERROR["store_name"],
                            "날짜": VLM_DRAFT_WITH_ERROR["date"],
                            "품목 수": len(VLM_DRAFT_WITH_ERROR["items"]),
                            "품목 합계": sum(
                                item["line_total"]
                                for item in VLM_DRAFT_WITH_ERROR["items"]
                            ),
                            "VLM 합계 초안": VLM_DRAFT_WITH_ERROR["total_amount"],
                            "합계 근거": VLM_DRAFT_WITH_ERROR["evidence"].get(
                                "total_amount"
                            ),
                        }}
                    ]
                )
            )
            display(pd.DataFrame(VLM_DRAFT_WITH_ERROR["items"]))
            """
        ),
        code(
            """
            # TODO(선택): VLM 초안을 보고 True 또는 False로 채우세요.
            MY_VLM_REVIEW = {
                "JSON 구조가 만들어졌다": None,
                "품목이 반복 행으로 정리되었다": None,
                "합계가 원본과 같다": None,
                "합계의 원본 근거가 있다": None,
                "지금 바로 Excel로 보내도 안전하다": None,
            }

            ANSWER_VLM_REVIEW = {
                "JSON 구조가 만들어졌다": True,
                "품목이 반복 행으로 정리되었다": True,
                "합계가 원본과 같다": False,
                "합계의 원본 근거가 있다": False,
                "지금 바로 Excel로 보내도 안전하다": False,
            }

            vlm_review_result = pd.DataFrame(
                [
                    {
                        "확인 질문": key,
                        "내 답": MY_VLM_REVIEW[key],
                        "전체 정답": expected,
                        "결과": (
                            "미작성"
                            if MY_VLM_REVIEW[key] is None
                            else "맞음"
                            if MY_VLM_REVIEW[key] == expected
                            else "다시 결과 확인"
                        ),
                    }
                    for key, expected in ANSWER_VLM_REVIEW.items()
                ]
            )
            display(vlm_review_result)
            """
        ),
        markdown(
            """
            ## 4. Document AI 검증으로 잘못된 합계를 찾습니다

            Document AI는 모델 이름 하나가 아니라 문서를 업무 데이터로 바꾸는
            기술과 과정입니다. 여기서는 VLM 초안에 다음 검사를 적용합니다.

            - 각 품목에서 `수량 × 단가 = 품목 금액`인가?
            - 품목 금액을 모두 더한 값과 합계가 같은가?
            - 합계가 원본에서 사람이 확인한 값과 같은가?
            - 합계가 나온 원본 문자열을 함께 저장했는가?
            """
        ),
        code(
            """
            EXPECTED_SOURCE_TOTAL = 76000
            EXPECTED_SOURCE_TEXT = "합계 금액 76,000"

            def validate_candidate(candidate):
                errors = []
                for index, item in enumerate(candidate["items"], start=1):
                    calculated = item["quantity"] * item["unit_price"]
                    if calculated != item["line_total"]:
                        errors.append(
                            f"{index}번째 품목: 수량×단가와 품목 금액이 다름"
                        )

                item_sum = sum(item["line_total"] for item in candidate["items"])
                if item_sum != candidate["total_amount"]:
                    errors.append(
                        f"품목 합계 {item_sum:,}원과 문서 합계 "
                        f"{candidate['total_amount']:,}원이 다름"
                    )
                if candidate["total_amount"] != EXPECTED_SOURCE_TOTAL:
                    errors.append(
                        f"원본 확인값 {EXPECTED_SOURCE_TOTAL:,}원과 문서 합계가 다름"
                    )

                evidence = candidate.get("evidence", {}).get("total_amount")
                if evidence != EXPECTED_SOURCE_TEXT:
                    errors.append("합계의 원본 근거가 없거나 원문과 다름")

                return {
                    "valid": not errors,
                    "item_sum": item_sum,
                    "errors": errors,
                    "next_action": (
                        "사람 원본 확인"
                        if not errors
                        else "값 수정 후 다시 검증"
                    ),
                }

            BEFORE_VALIDATION = validate_candidate(VLM_DRAFT_WITH_ERROR)
            assert BEFORE_VALIDATION["valid"] is False
            display(pd.DataFrame({"발견한 문제": BEFORE_VALIDATION["errors"]}))
            """
        ),
        code(
            """
            # TODO(선택): 원본을 보고 두 값을 채운 뒤 이 셀을 다시 실행하세요.
            MY_CORRECTED_TOTAL = None
            MY_TOTAL_EVIDENCE = None

            learner_candidate = copy.deepcopy(VLM_DRAFT_WITH_ERROR)
            if MY_CORRECTED_TOTAL is not None:
                learner_candidate["total_amount"] = MY_CORRECTED_TOTAL
            if MY_TOTAL_EVIDENCE is not None:
                learner_candidate["evidence"]["total_amount"] = MY_TOTAL_EVIDENCE

            LEARNER_VALIDATION = validate_candidate(learner_candidate)
            display(
                pd.DataFrame(
                    [
                        {
                            "수정한 합계": learner_candidate["total_amount"],
                            "원본 근거": learner_candidate["evidence"].get(
                                "total_amount"
                            ),
                            "검증 통과": LEARNER_VALIDATION["valid"],
                            "다음 행동": LEARNER_VALIDATION["next_action"],
                        }
                    ]
                )
            )
            if LEARNER_VALIDATION["errors"]:
                display(pd.DataFrame({"남은 문제": LEARNER_VALIDATION["errors"]}))
            """
        ),
        markdown(
            """
            <details>
            <summary>수정 힌트와 전체 정답 보기</summary>

            - 합계는 정수 `76000`입니다.
            - 근거에는 원본의 줄 전체인 `합계 금액 76,000`을 적습니다.
            </details>
            """
        ),
        code(
            """
            ANSWER_CORRECTED_TOTAL = 76000
            ANSWER_TOTAL_EVIDENCE = "합계 금액 76,000"

            VALIDATED_RECEIPT = copy.deepcopy(VLM_DRAFT_WITH_ERROR)
            VALIDATED_RECEIPT["total_amount"] = ANSWER_CORRECTED_TOTAL
            VALIDATED_RECEIPT["evidence"]["total_amount"] = ANSWER_TOTAL_EVIDENCE
            AFTER_VALIDATION = validate_candidate(VALIDATED_RECEIPT)

            assert AFTER_VALIDATION["valid"] is True
            display(
                pd.DataFrame(
                    [
                        {
                            "수정 전": VLM_DRAFT_WITH_ERROR["total_amount"],
                            "수정 후": VALIDATED_RECEIPT["total_amount"],
                            "품목 합계": AFTER_VALIDATION["item_sum"],
                            "원본 근거": VALIDATED_RECEIPT["evidence"]["total_amount"],
                            "검증 결과": "통과",
                        }
                    ]
                )
            )
            """
        ),
        markdown(
            """
            ## 5. 검증 통과와 사람 승인을 구분합니다

            계산 규칙을 통과해도 Excel이 바로 만들어지지는 않습니다. 담당자가
            원본과 수정 결과를 비교하고 승인해야 합니다. 예외 처리, 사람 승인,
            업무 시스템 연결까지 운영하는 범위를 IDP라고 설명할 수 있습니다.
            """
        ),
        code(
            """
            # TODO(선택): 원본 확인을 마쳤다면 "APPROVED", 아니면 "PENDING"을 적으세요.
            MY_REVIEW_DECISION = None

            if not AFTER_VALIDATION["valid"]:
                learner_export_status = "검증 오류로 저장 차단"
            elif MY_REVIEW_DECISION != "APPROVED":
                learner_export_status = "사람 승인 전이므로 저장 대기"
            else:
                learner_export_status = "Excel 저장 가능"

            print("내 결정:", MY_REVIEW_DECISION)
            print("현재 상태:", learner_export_status)

            # 전체 정답: 공개 원본을 대조한 뒤 승인한 경우
            ANSWER_REVIEW_DECISION = "APPROVED"
            EXCEL_PREVIEW = pd.DataFrame(
                [
                    {
                        "상호명": VALIDATED_RECEIPT["store_name"],
                        "날짜": VALIDATED_RECEIPT["date"],
                        "품목 수": len(VALIDATED_RECEIPT["items"]),
                        "합계": VALIDATED_RECEIPT["total_amount"],
                        "검토 결정": ANSWER_REVIEW_DECISION,
                    }
                ]
            )
            display(EXCEL_PREVIEW)
            """
        ),
        markdown(
            """
            ## 6. 어떤 기술의 역할인지 직접 구분합니다

            다음 상황에서 가장 중심이 되는 개념을 적어 보세요.

            - `멀티모달 AI`는 이미지·텍스트처럼 여러 정보 형식을 함께 다루는 상위 범주입니다.
            - `VLM`은 그중 이미지와 언어를 함께 처리하는 모델입니다.
            - OCR과 VLM은 필요에 따라 선택하거나 조합할 수 있으며 고정 순서가 아닙니다.
            """
        ),
        code(
            """
            # TODO: OCR, 멀티모달 AI, VLM, Document AI, IDP 중에서 고르세요.
            MY_CONCEPT_CHOICES = {
                "글자의 문자열·위치·신뢰도를 얻는다": None,
                "이미지와 텍스트 등 여러 형식을 함께 다룬다": None,
                "문서 이미지와 지시문으로 JSON 초안을 만든다": None,
                "스키마·원본 근거·업무 규칙으로 결과를 검사한다": None,
                "예외·사람 승인·Excel 연결을 운영한다": None,
            }

            ANSWER_CONCEPT_CHOICES = {
                "글자의 문자열·위치·신뢰도를 얻는다": "OCR",
                "이미지와 텍스트 등 여러 형식을 함께 다룬다": "멀티모달 AI",
                "문서 이미지와 지시문으로 JSON 초안을 만든다": "VLM",
                "스키마·원본 근거·업무 규칙으로 결과를 검사한다": "Document AI",
                "예외·사람 승인·Excel 연결을 운영한다": "IDP",
            }

            concept_quiz = pd.DataFrame(
                [
                    {
                        "상황": situation,
                        "내 답": MY_CONCEPT_CHOICES[situation],
                        "전체 정답": answer,
                        "결과": (
                            "미작성"
                            if MY_CONCEPT_CHOICES[situation] is None
                            else "맞음"
                            if MY_CONCEPT_CHOICES[situation] == answer
                            else "다시 비교"
                        ),
                    }
                    for situation, answer in ANSWER_CONCEPT_CHOICES.items()
                ]
            )
            display(concept_quiz)
            """
        ),
        markdown(
            """
            ## 한눈에 정리

            | 개념 | 이번 영수증에서 본 역할 | 결과를 바로 믿으면 안 되는 이유 |
            | --- | --- | --- |
            | OCR | 글자·위치·신뢰도 판독 | 높은 신뢰도에서도 글자가 틀릴 수 있음 |
            | 멀티모달 AI | 이미지와 텍스트를 함께 다루는 상위 범주 | 특정 처리 단계나 제품 이름이 아님 |
            | VLM | 문맥을 이용해 JSON 구조 초안 생성 | 그럴듯한 잘못된 값을 만들 수 있음 |
            | Document AI | 추출·정규화·근거·업무 규칙 검증 | 규칙이 놓치는 오류는 사람이 확인해야 함 |
            | IDP | 예외 처리·사람 승인·Excel 연결·운영 개선 | 모델만으로 회사 업무가 완성되지 않음 |

            2교시에서는 이 가운데 OCR을 직접 실행해 글자·위치·신뢰도를 만들어 봅니다.
            """
        ),
        code(
            """
            comparison_report = {
                "source_document": "taebaek_restaurant_2025_redacted.png",
                "learner_attempts": {
                    "source_observation": MY_SOURCE_OBSERVATION,
                    "ocr_review": MY_OCR_REVIEW,
                    "vlm_review": MY_VLM_REVIEW,
                    "corrected_total": MY_CORRECTED_TOTAL,
                    "total_evidence": MY_TOTAL_EVIDENCE,
                    "validation": LEARNER_VALIDATION,
                    "review_decision": MY_REVIEW_DECISION,
                    "concept_choices": MY_CONCEPT_CHOICES,
                },
                "source_observation": ANSWER_SOURCE_OBSERVATION,
                "ocr": {
                    "engine": "PP-OCRv5 실제 실행 기록",
                    "token_count": len(RECORDED_PP_OCRV5_TOKENS),
                    "review_answer": ANSWER_OCR_REVIEW,
                },
                "vlm": {
                    "example": "교육용 오류 삽입 구조 초안",
                    "model_called_in_this_notebook": False,
                    "review_answer": ANSWER_VLM_REVIEW,
                },
                "document_ai": {
                    "before_validation": BEFORE_VALIDATION,
                    "after_validation": AFTER_VALIDATION,
                    "corrected_total": VALIDATED_RECEIPT["total_amount"],
                    "source_evidence": VALIDATED_RECEIPT["evidence"]["total_amount"],
                },
                "idp": {
                    "human_review_decision": ANSWER_REVIEW_DECISION,
                    "export_rule": "검증 통과와 사람 승인 뒤 Excel 저장",
                },
                "concept_answers": ANSWER_CONCEPT_CHOICES,
            }

            output_path = OUTPUT_DIR / "lesson01_comparison_report.json"
            output_path.write_text(
                json.dumps(comparison_report, ensure_ascii=False, indent=2) + "\\n",
                encoding="utf-8",
            )

            assert comparison_report["ocr"]["token_count"] == 44
            assert comparison_report["document_ai"]["before_validation"]["valid"] is False
            assert comparison_report["document_ai"]["after_validation"]["valid"] is True
            assert comparison_report["idp"]["human_review_decision"] == "APPROVED"
            print("CHECKPOINT 1/1 PASS:", output_path)
            """
        ),
    ]
    return notebook("01_document_ai_overview.ipynb", cells)


def notebook_02() -> dict:
    cells = [
        intro(
            2,
            "OCR 기반 텍스트 추출 실습",
            "ocr_result.json",
            "공개 한국 영수증에 실제 OCR을 실행하고 원본 위 좌표·신뢰도를 확인합니다.",
        ),
        runtime_cell(),
        code(
            """
            import hashlib
            import io
            import pandas as pd
            from PIL import Image, ImageDraw
            try:
                from IPython.display import display
            except ImportError:
                def display(value):
                    print(value)

            SAMPLE_IMAGE_PATH = (
                "sample_docs/public_receipts/korea/"
                "taebaek_restaurant_2025_redacted.png"
            )
            RECORDED_OCR_PATH = (
                "tests/fixtures/ppocrv5_recorded_receipt_tokens.json"
            )
            OCR_METADATA_PATH = (
                "tests/fixtures/ppocrv5_recorded_receipt_metadata.json"
            )
            lesson_assets = load_course_assets(
                SAMPLE_IMAGE_PATH,
                RECORDED_OCR_PATH,
                OCR_METADATA_PATH,
            )
            course_example_image = Image.open(
                io.BytesIO(lesson_assets[SAMPLE_IMAGE_PATH])
            ).convert("RGB")
            receipt_image = course_example_image.copy()
            COURSE_EXAMPLE_OCR_RESULT = json.loads(
                lesson_assets[RECORDED_OCR_PATH].decode("utf-8")
            )
            OCR_RECORD_METADATA = json.loads(
                lesson_assets[OCR_METADATA_PATH].decode("utf-8")
            )
            assert hashlib.sha256(
                lesson_assets[SAMPLE_IMAGE_PATH]
            ).hexdigest() == OCR_RECORD_METADATA["source_image_sha256"]
            assert hashlib.sha256(
                lesson_assets[RECORDED_OCR_PATH]
            ).hexdigest() == OCR_RECORD_METADATA["token_file_sha256"]
            assert course_example_image.size == (
                OCR_RECORD_METADATA["source_image_size"]["width"],
                OCR_RECORD_METADATA["source_image_size"]["height"],
            )
            for item in COURSE_EXAMPLE_OCR_RESULT:
                item["confidence_source"] = "previous_actual_ocr_run"

            # TODO(선택): 내 사진이나 다른 공개 이미지를 시험할 때만 True로 바꾸세요.
            USE_MY_FILE = False
            INPUT_FILE_NAME = "taebaek_restaurant_2025_redacted.png"
            if USE_MY_FILE and not VALIDATION_MODE:
                from google.colab import files
                print(
                    "JPG·JPEG·PNG·WEBP 사진 한 장을 선택하세요. "
                    "인터넷 자료는 이용조건을 확인하고, 내 문서는 카드·전화·"
                    "회원번호 등 식별정보를 먼저 가립니다."
                )
                uploaded = files.upload()
                if len(uploaded) != 1:
                    raise ValueError("문서 이미지 한 장만 선택하세요.")
                INPUT_FILE_NAME, uploaded_bytes = next(iter(uploaded.items()))
                if Path(INPUT_FILE_NAME).suffix.lower() not in {
                    ".jpg", ".jpeg", ".png", ".webp"
                }:
                    raise ValueError(
                        "2교시는 JPG·JPEG·PNG·WEBP 한 장을 사용합니다. "
                        "PDF·Office 문서는 8교시에서 체험합니다."
                    )
                receipt_image = Image.open(
                    io.BytesIO(uploaded_bytes)
                ).convert("RGB")
            RUN_OCR_NOW = not VALIDATION_MODE
            print(
                "처리 계획:",
                (
                    "지금 이 사진을 OCR로 직접 읽습니다."
                    if RUN_OCR_NOW
                    else "자동검사에서는 수업용 예제 결과를 사용합니다."
                ),
            )
            print("입력 파일:", INPUT_FILE_NAME)
            """
        ),
        markdown(
            """
            ## 실행

            **PaddleOCR**는 실행 도구이고, **PP-OCRv5 Korean**은 그 안에서
            사용하는 한국어 OCR 모델입니다. 이 노트북은 PaddleOCR 3.7에서
            현재 이미지 한 장을 실제로 읽습니다.

            설치·모델 다운로드가 3분을 넘기면 중지합니다. 오류 메시지를 보존한
            채 같은 공개 영수증을 이전에 실제로 읽어 보존한 결과로 전환합니다.
            이때 현재 사진을 분석한 결과가 아니라는 안내가 표시됩니다.
            """
        ),
        code(
            """
            OCR_RESULT = COURSE_EXAMPLE_OCR_RESULT
            PROCESSING_PATH = "COURSE_EXAMPLE"
            OCR_ERROR = "자동검사용 실행"

            if RUN_OCR_NOW:
                import subprocess
                try:
                    subprocess.check_call(
                        [sys.executable, "-m", "pip", "install", "-q",
                         "paddlepaddle==3.2.1", "paddleocr==3.7.0"]
                    )
                    from paddleocr import PaddleOCR

                    image_path = OUTPUT_DIR / "golden_receipt.jpg"
                    receipt_image.save(image_path)
                    engine = PaddleOCR(
                        lang="korean",
                        ocr_version="PP-OCRv5",
                        use_doc_orientation_classify=False,
                        use_doc_unwarping=False,
                        use_textline_orientation=False,
                        device="cpu",
                    )
                    page = list(engine.predict(str(image_path)))[0]
                    payload = page.json() if callable(page.json) else page.json
                    result = payload.get("res", payload)
                    OCR_RESULT = [
                        {
                            "box": box.tolist() if hasattr(box, "tolist") else box,
                            "text": text,
                            "confidence": float(score),
                        }
                        for box, text, score in zip(
                            result.get("rec_polys", []),
                            result.get("rec_texts", []),
                            result.get("rec_scores", []),
                        )
                    ]
                    PROCESSING_PATH = "DIRECT_OCR"
                    OCR_ERROR = ""
                except Exception as exc:
                    PROCESSING_PATH = "COURSE_EXAMPLE"
                    OCR_ERROR = f"{type(exc).__name__}: {exc}"

            if PROCESSING_PATH == "DIRECT_OCR":
                print("처리 방식: 지금 이 사진을 OCR로 직접 읽었습니다.")
            else:
                print("처리 방식: 이전 실제 OCR 기록을 불러왔습니다.")
                print("중요: 현재 사진을 분석한 결과가 아닙니다.")
            if OCR_ERROR:
                print("직접 읽지 못한 이유:", OCR_ERROR)
            if PROCESSING_PATH == "COURSE_EXAMPLE":
                if INPUT_FILE_NAME != "taebaek_restaurant_2025_redacted.png":
                    print(
                        "내 영수증을 직접 읽지 못해 박스 표시는 "
                        "공개 영수증의 이전 실제 OCR 기록으로 바꿉니다."
                    )
                receipt_image = course_example_image.copy()
                DISPLAY_INPUT_FILE_NAME = "taebaek_restaurant_2025_redacted.png"
                OCR_COORDINATE_SIZE = (
                    OCR_RECORD_METADATA["coordinate_space"]["width"],
                    OCR_RECORD_METADATA["coordinate_space"]["height"],
                )
                expected_height = round(
                    receipt_image.height
                    * OCR_COORDINATE_SIZE[0]
                    / receipt_image.width
                )
                assert OCR_COORDINATE_SIZE[1] == expected_height
            else:
                DISPLAY_INPUT_FILE_NAME = INPUT_FILE_NAME
                OCR_COORDINATE_SIZE = receipt_image.size
            print("판독 영역:", len(OCR_RESULT))
            """
        ),
        code(
            """
            annotated = receipt_image.copy()
            draw = ImageDraw.Draw(annotated)
            scale_x = annotated.width / OCR_COORDINATE_SIZE[0]
            scale_y = annotated.height / OCR_COORDINATE_SIZE[1]
            for item in OCR_RESULT:
                points = item["box"]
                xs = [point[0] * scale_x for point in points]
                ys = [point[1] * scale_y for point in points]
                draw.rectangle(
                    (min(xs), min(ys), max(xs), max(ys)),
                    outline="#0F766E",
                    width=4,
                )
            annotated_path = OUTPUT_DIR / "ocr_boxes.png"
            annotated.save(annotated_path)

            annotated_preview = annotated.copy()
            annotated_preview.thumbnail((650, 800))
            print("1) 원본 위 OCR 탐지 영역")
            display(annotated_preview)

            result_table = pd.DataFrame(
                [
                    {
                        "OCR 글자": item["text"],
                        "신뢰도": round(float(item["confidence"] or 0), 3),
                    }
                    for item in OCR_RESULT
                    if item.get("text")
                ]
            )
            print("2) 인식한 글자와 신뢰도")
            display(result_table)
            print("3) 확인할 곳: 상자가 글자를 감싸는지, 낮은 신뢰도 글자가 틀렸는지")

            output = {
                "processing_path": PROCESSING_PATH,
                "ocr_error": OCR_ERROR,
                "input_file": DISPLAY_INPUT_FILE_NAME,
                "image_size": {
                    "width": annotated.width,
                    "height": annotated.height,
                },
                "ocr_coordinate_size": {
                    "width": OCR_COORDINATE_SIZE[0],
                    "height": OCR_COORDINATE_SIZE[1],
                },
                "items": [
                    {**item, "matches_source": None, "review_note": ""}
                    for item in OCR_RESULT
                ],
            }
            output_path = OUTPUT_DIR / "ocr_result.json"
            output_path.write_text(
                json.dumps(output, ensure_ascii=False, indent=2) + "\\n",
                encoding="utf-8",
            )
            checkpoint_label = (
                "직접 OCR 실행"
                if PROCESSING_PATH == "DIRECT_OCR"
                else "이전 실제 OCR 기록 사용"
            )
            print(
                "CHECKPOINT 1/1 PASS:",
                checkpoint_label,
                output_path,
                annotated_path,
            )
            """
        ),
        markdown(
            """
            ## 내가 직접 채우는 3줄

            금액·날짜처럼 원본 대조가 필요한 OCR 토큰을 키워드로 표시합니다.
            """
        ),
        code(
            """
            # TODO: 원본 대조할 키워드 세 개를 넣으세요.
            review_keywords = [None, None, None]
            if any(value is None for value in review_keywords):
                print("빈칸이 있습니다. 아래 전체 정답과 비교하세요.")
            """
        ),
        markdown(
            """
            <details>
            <summary>힌트와 전체 정답 보기</summary>

            날짜의 연도, 합계 라벨, 합계 금액처럼 영향이 큰 토큰을 고릅니다.
            </details>
            """
        ),
        code(
            """
            ANSWER_REVIEW_KEYWORDS = ["이태리", "2025", "76,000"]
            marked = 0
            for item in output["items"]:
                if any(
                    keyword in item.get("text", "")
                    for keyword in ANSWER_REVIEW_KEYWORDS
                ):
                    item["review_note"] = "원본 대조 필수"
                    marked += 1
            output_path.write_text(
                json.dumps(output, ensure_ascii=False, indent=2) + "\\n",
                encoding="utf-8",
            )
            assert marked == 3
            print("전체 정답 · 원본 대조 표시:", marked, "개")
            import zipfile
            bundle_path = OUTPUT_DIR / "lesson02_ocr_outputs.zip"
            with zipfile.ZipFile(bundle_path, "w") as archive:
                archive.write(output_path, output_path.name)
                archive.write(annotated_path, annotated_path.name)
            print("한 번만 다운로드할 묶음:", bundle_path)
            download_artifact(bundle_path)
            """
        ),
    ]
    return notebook("02_ocr_basic.ipynb", cells)


def notebook_03() -> dict:
    cells = [
        intro(
            3,
            "문서 구조 이해 및 추출 결과 정제",
            "clean_receipt.json",
            "2교시 결과를 불러와 원문은 보존하고, 공백·날짜·표 영역만 정리합니다.",
        ),
        runtime_cell(),
        code(golden_constants()),
        code(spatial_reconstruction_source()),
        code(
            """
            import re

            previous_path = OUTPUT_DIR / "ocr_result.json"
            # 기본값 True: 새 Colab에서도 공개 준비 입력으로 바로 실행합니다.
            # 앞 교시 파일을 이어 쓰려면 False로 바꾸고 업로드 창에서 선택합니다.
            USE_COURSE_EXAMPLE = True
            if not previous_path.exists() and not USE_COURSE_EXAMPLE:
                upload_previous_artifact("ocr_result.json")
            if previous_path.exists():
                previous = json.loads(previous_path.read_text(encoding="utf-8"))
                raw_text = "\\n".join(item["text"] for item in previous["items"])
                layout_lines = reconstruct_spatial_lines(previous["items"])
                INPUT_MODE = "PREVIOUS_LESSON"
            else:
                raw_text = GOLDEN_OCR_TEXT
                layout_lines = raw_text.splitlines()
                INPUT_MODE = "COURSE_EXAMPLE"
            print(
                "입력 자료:",
                (
                    "2교시에서 만든 OCR 결과를 불러왔습니다."
                    if INPUT_MODE == "PREVIOUS_LESSON"
                    else "수업용 예제 OCR 결과를 불러왔습니다."
                ),
            )
            if INPUT_MODE == "COURSE_EXAMPLE":
                print("중요: 지금 새로 OCR을 실행한 결과가 아닙니다.")


            def clean_lines(text, source_lines):
                cleaned = []
                changes = []
                for raw in source_lines:
                    normalized = re.sub(r"\\s+", " ", raw.strip())
                    if normalized:
                        cleaned.append(normalized)
                    if raw != normalized:
                        changes.append({"before": raw, "after": normalized})
                groups = {"header": [], "date": [], "items": [], "total": [], "other": []}
                for line in cleaned:
                    if re.search(r"\\d{4}[-./]\\d{1,2}[-./]\\d{1,2}", line):
                        groups["date"].append(line)
                    elif "합계" in line:
                        groups["total"].append(line)
                    elif re.search(r"[\\d,]+\\s+\\d+\\s+[\\d,]+$", line):
                        groups["items"].append(line)
                    elif not groups["header"]:
                        groups["header"].append(line)
                    else:
                        groups["other"].append(line)
                return {
                    "input_mode": INPUT_MODE,
                    "raw_text": text,
                    "layout_lines": source_lines,
                    "cleaned_lines": cleaned,
                    "groups": groups,
                    "change_log": changes,
                    "rule": "원문에 없는 값은 추가하지 않음",
                }


            clean_result = clean_lines(raw_text, layout_lines)
            output_path = OUTPUT_DIR / "clean_receipt.json"
            output_path.write_text(
                json.dumps(clean_result, ensure_ascii=False, indent=2) + "\\n",
                encoding="utf-8",
            )
            assert clean_result["raw_text"] == raw_text
            print("원문 줄:", len(raw_text.splitlines()))
            print("품목 후보 줄:", len(clean_result["groups"]["items"]))
            print("CHECKPOINT 1/1 PASS:", output_path)
            download_artifact(output_path)
            """
        ),
        markdown(
            """
            ## 내가 직접 채우는 1줄

            품목 행은 끝부분에 `단가 수량 금액`이 반복됩니다. 그 모양을 찾는
            정규식 한 줄을 채웁니다.
            """
        ),
        code(
            """
            # TODO: 품목 행 끝의 숫자 세 묶음을 찾는 정규식을 넣으세요.
            my_item_rule = None
            if my_item_rule is None:
                print("빈칸입니다. 아래 전체 정답과 비교하세요.")
            """
        ),
        markdown(
            """
            <details>
            <summary>힌트와 전체 정답 보기</summary>

            숫자·쉼표 묶음, 수량 정수, 마지막 숫자·쉼표 묶음을 공백으로 연결합니다.
            </details>
            """
        ),
        code(
            r"""
            ANSWER_ITEM_RULE = r"[\d,]+\s+\d+\s+[\d,]+$"
            answer_item_lines = [
                line
                for line in clean_result["cleaned_lines"]
                if re.search(ANSWER_ITEM_RULE, line)
            ]
            assert answer_item_lines
            print("전체 정답 · 찾은 품목 후보:", answer_item_lines)
            """
        ),
    ]
    return notebook("03_document_structure.ipynb", cells)


def notebook_04() -> dict:
    cells = [
        intro(
            4,
            "멀티모달·생성형 AI 기반 핵심 정보 추출",
            "receipt.json",
            "같은 영수증을 업무 JSON 초안으로 만들고 모든 핵심값에 원본 근거를 붙입니다.",
        ),
        runtime_cell(),
        code(golden_constants()),
        code(parser_source()),
        markdown(
            """
            ## OCR+규칙과 VLM 구조 초안은 다른 경로입니다

            이 교시에서는 두 결과를 나란히 봅니다.

            - **내 문서 경로**: 3교시 OCR 결과에 규칙 추출을 적용합니다.
            - **VLM 비교 경로**: 같은 공개 영수증을 표 Markdown으로 구조화한
              수업용 VLM 구조 예제를 사용합니다.

            비교 예제는 지금 모델을 실행한 결과가 아닙니다. 강사의 VLM 직접 시연 또는
            녹화가 실제 호출 경험을 담당하며, 필수 실습에서는 비용·GPU·계정
            변수를 없앱니다. 어느 경로든 다음 세 가지를 확인합니다.

            1. **스키마**: 필요한 필드와 자료형이 맞는가?
            2. **근거**: 값이 원본 어느 줄에서 왔는가?
            3. **불확실성**: 근거가 없으면 추측하지 않고 `null`인가?
            """
        ),
        code(
            """
            previous_path = OUTPUT_DIR / "clean_receipt.json"
            # 기본값 True: 새 Colab에서도 공개 준비 입력으로 바로 실행합니다.
            # 앞 교시 파일을 이어 쓰려면 False로 바꾸고 업로드 창에서 선택합니다.
            USE_COURSE_EXAMPLE = True
            if not previous_path.exists() and not USE_COURSE_EXAMPLE:
                upload_previous_artifact("clean_receipt.json")
            if previous_path.exists():
                clean_result = json.loads(previous_path.read_text(encoding="utf-8"))
                source_text = "\\n".join(clean_result["cleaned_lines"])
                INPUT_MODE = "PREVIOUS_LESSON"
            else:
                source_text = GOLDEN_OCR_TEXT
                INPUT_MODE = "COURSE_EXAMPLE"

            receipt_source_mode = (
                "ocr_rule_extraction_from_previous_lesson"
                if INPUT_MODE == "PREVIOUS_LESSON"
                else "course_example_rule_extraction"
            )
            receipt = extract_receipt_from_text(
                source_text,
                receipt_source_mode,
            )
            receipt["provenance"] = {
                "fixture_type": (
                    "previous_lesson_artifact"
                    if INPUT_MODE == "PREVIOUS_LESSON"
                    else "human_verified_transcription_fixture"
                ),
                "input_file": (
                    "clean_receipt.json"
                    if INPUT_MODE == "PREVIOUS_LESSON"
                    else "course example GOLDEN_OCR_TEXT"
                ),
                "engine": "course_rule_extractor",
                "engine_version": "v2",
                "target_technology": "OCR + rule baseline",
                "recorded_at": "2026-07-28",
                "reviewer": (
                    "learner"
                    if INPUT_MODE == "PREVIOUS_LESSON"
                    else "교육자료 검수자"
                ),
                "disclaimer": "이 receipt.json은 VLM 결과가 아니라 OCR+규칙 기준선입니다.",
            }
            receipt["input_mode"] = INPUT_MODE
            receipt["source_text"] = source_text

            vlm_demo = extract_receipt_from_text(
                GOLDEN_VLM_MARKDOWN,
                "course_example_vlm_structure_rule_extraction",
            )
            vlm_demo["provenance"] = {
                "fixture_type": "course_example",
                "input_file": "taebaek_restaurant_2025_redacted.png",
                "engine": "not_executed",
                "engine_version": "not_applicable",
                "target_technology": "PaddleOCR-VL-1.6",
                "recorded_at": "2026-07-28",
                "reviewer": "교육자료 검수자",
                "disclaimer": "현재 실행에서 VLM을 호출한 결과가 아닙니다.",
            }

            comparison = {
                field: {
                    "ocr_rule": receipt.get(field),
                    "course_example_vlm_structure": vlm_demo.get(field),
                    "must_check_source": True,
                }
                for field in ("store_name", "date", "total_amount", "items")
            }
            comparison_path = OUTPUT_DIR / "vlm_comparison.json"
            comparison_path.write_text(
                json.dumps({
                    "warning": "수업용 VLM 구조 예제이며 지금 모델을 실행한 결과가 아님",
                    "comparison": comparison,
                    "vlm_provenance": vlm_demo["provenance"],
                }, ensure_ascii=False, indent=2) + "\\n",
                encoding="utf-8",
            )

            assert receipt["total_amount"] is None or isinstance(
                receipt["total_amount"], int
            )
            if INPUT_MODE == "COURSE_EXAMPLE":
                assert receipt["total_amount"] == 76000
            assert vlm_demo["total_amount"] == 76000
            assert len(vlm_demo["items"]) == 5
            output_path = OUTPUT_DIR / "receipt.json"
            output_path.write_text(
                json.dumps(receipt, ensure_ascii=False, indent=2) + "\\n",
                encoding="utf-8",
            )
            print(json.dumps({
                "total_amount": receipt["total_amount"],
                "evidence": receipt["evidence"]["total_amount"],
                "source_mode": receipt["source_mode"],
                "vlm_demo_mode": vlm_demo["source_mode"],
            }, ensure_ascii=False, indent=2))
            print("CHECKPOINT 1/1 PASS:", output_path, comparison_path)
            download_artifact(output_path)
            download_artifact(comparison_path)
            """
        ),
        markdown(
            """
            ## 내가 직접 채우는 5줄

            아래 셀에서 원본 대조가 가장 중요한 필드 하나와 처리 결정을
            입력합니다. 막히면 바로 다음 정답 셀을 열어 비교합니다.
            """
        ),
        code(
            """
            # TODO: None 세 곳을 채우세요.
            my_review = {
                "field": None,
                "evidence_found": None,
                "action": None,
            }
            if None in my_review.values():
                print("빈칸이 있습니다. 아래 힌트·정답 셀과 비교하세요.")
            else:
                print("내 검토 결정:", my_review)
            """
        ),
        markdown(
            """
            <details>
            <summary>힌트와 전체 정답 보기</summary>

            영향이 큰 `total_amount`를 선택하고, 원본 근거가 있으면
            `REVIEW_BEFORE_EXPORT`로 둡니다.
            </details>
            """
        ),
        code(
            """
            ANSWER_REVIEW = {
                "field": "total_amount",
                "evidence_found": bool(receipt["evidence"]["total_amount"]["raw_value"]),
                "action": (
                    "REVIEW_BEFORE_EXPORT"
                    if receipt["evidence"]["total_amount"]["raw_value"]
                    else "MANUAL_REVIEW_REQUIRED"
                ),
            }
            assert ANSWER_REVIEW["action"] in {
                "REVIEW_BEFORE_EXPORT",
                "MANUAL_REVIEW_REQUIRED",
            }
            print("전체 정답:", ANSWER_REVIEW)
            """
        ),
    ]
    return notebook("04_genai_extraction.ipynb", cells)


STREAMLIT_SETUP = """
import importlib.metadata
import subprocess

required_streamlit = "1.60.0"
try:
    installed_streamlit = importlib.metadata.version("streamlit")
except importlib.metadata.PackageNotFoundError:
    installed_streamlit = None
if installed_streamlit != required_streamlit:
    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", "-q", f"streamlit=={required_streamlit}"]
    )
"""

PADDLEOCR_SETUP = """
required_paddlepaddle = "3.2.1"
required_paddleocr = "3.7.0"
if VALIDATION_MODE:
    PADDLEOCR_READY = False
    print("자동검증 모드: PaddleOCR 설치를 생략합니다.")
else:
    try:
        installed_paddlepaddle = importlib.metadata.version("paddlepaddle")
        installed_paddleocr = importlib.metadata.version("paddleocr")
        if (
            installed_paddlepaddle != required_paddlepaddle
            or installed_paddleocr != required_paddleocr
        ):
            raise importlib.metadata.PackageNotFoundError
        PADDLEOCR_READY = True
    except importlib.metadata.PackageNotFoundError:
        try:
            subprocess.check_call(
                [
                    sys.executable,
                    "-m",
                    "pip",
                    "install",
                    "-q",
                    f"paddlepaddle=={required_paddlepaddle}",
                    f"paddleocr=={required_paddleocr}",
                ]
            )
            PADDLEOCR_READY = True
        except Exception as exc:
            PADDLEOCR_READY = False
            print("영수증 직접 읽기 준비 실패:", type(exc).__name__, exc)

print("Streamlit:", importlib.metadata.version("streamlit"))
print(
    "영수증 직접 읽기:",
    (
        f"준비 완료 · PaddleOCR {importlib.metadata.version('paddleocr')}"
        if PADDLEOCR_READY
        else "준비하지 못했습니다 · 수업용 예제로 계속할 수 있습니다."
    ),
)
"""


def notebook_05() -> dict:
    app_source = (
        "import streamlit as st\n\n"
        f"GOLDEN_RECEIPT = "
        f"{pformat(GOLDEN_RECEIPT, sort_dicts=False, width=88)}\n"
        f"GOLDEN_OCR_TEXT = {GOLDEN_OCR_TEXT!r}\n\n"
        + dedent(
            f'''
            st.set_page_config(page_title="영수증 Document AI", layout="wide")
            st.title("영수증 Document AI 미니 앱")
            uploaded = st.file_uploader(
                "승인된 비식별 이미지 또는 PDF 한 장 · 최대 5MB",
                type=["png", "jpg", "jpeg", "pdf"],
                max_upload_size=5,
                help="PNG, JPG, JPEG, PDF만 허용합니다. 수업에서는 한 번에 5MB 이하 한 장만 처리합니다.",
            )
            if uploaded is not None:
                st.success(f"업로드 연결 확인: {{uploaded.name}} · {{len(uploaded.getvalue()):,}} bytes")
                st.caption("이 파일은 6교시에서 실제 처리 함수와 연결합니다.")

            if st.button("수업용 예제 결과 보기"):
                st.info(
                    "수업용 예제 결과를 불러왔습니다. "
                    "현재 업로드한 파일을 분석한 결과가 아닙니다."
                )
                st.text_area("판독 원문", GOLDEN_OCR_TEXT, height=220)
                st.json(GOLDEN_RECEIPT)
            '''
        ).lstrip()
    )
    app_assignment = readable_string_assignment("app_code", app_source)
    cells = [
        intro(
            5,
            "문서 자동화 웹 애플리케이션 기본 구현",
            "app_05.py",
            "업로드·실행·원문·JSON 영역을 만들고 업로드한 파일명이 화면에 반영되는지 확인합니다.",
        ),
        runtime_cell(),
        code(STREAMLIT_SETUP),
        code(
            app_assignment
            + """

output_path = OUTPUT_DIR / "app_05.py"
output_path.write_text(app_code, encoding="utf-8")
print("저장:", output_path)
"""
        ),
        markdown(
            """
            ## 내가 직접 바꾸는 화면 문구 2개

            앱 제목과 실행 버튼 문구를 업무 사용자가 이해할 표현으로 바꿉니다.
            """
        ),
        code(
            """
            # TODO: None 두 곳을 채우세요.
            my_app_title = None
            my_button_label = None
            if None in (my_app_title, my_button_label):
                print("빈칸이 있습니다. 아래 전체 정답과 비교하세요.")
            """
        ),
        markdown(
            """
            <details>
            <summary>힌트와 전체 정답 보기</summary>

            문서 종류와 버튼을 눌렀을 때 일어나는 일을 그대로 적습니다.
            </details>
            """
        ),
        code(
            """
            ANSWER_APP_TITLE = my_app_title or "영수증 검토 미니 앱"
            ANSWER_BUTTON_LABEL = my_button_label or "공개 영수증 결과 확인"
            app_code = app_code.replace(
                "영수증 Document AI 미니 앱",
                ANSWER_APP_TITLE,
            ).replace(
                "수업용 예제 결과 보기",
                ANSWER_BUTTON_LABEL,
            )
            output_path.write_text(app_code, encoding="utf-8")
            print("내 화면 문구:", ANSWER_APP_TITLE, "/", ANSWER_BUTTON_LABEL)
            """
        ),
        code(
            """
            from streamlit.testing.v1 import AppTest

            app_test = AppTest.from_file(str(output_path)).run(timeout=20)
            assert not app_test.exception
            assert app_test.title[0].value == ANSWER_APP_TITLE
            assert len(app_test.file_uploader) == 1
            assert len(app_test.button) == 1
            app_test.button[0].click().run(timeout=20)
            assert any(
                "현재 업로드한 파일을 분석한 결과가 아닙니다" in item.value
                for item in app_test.info
            )
            print("CHECKPOINT 1/1 PASS: 업로드·버튼·결과 화면")
            """
        ),
        streamlit_preview_cell("output_path", 8505),
    ]
    return notebook("05_streamlit_basic.ipynb", cells)


def notebook_06() -> dict:
    app_source = (
        "import tempfile\n"
        "from pathlib import Path\n"
        "import streamlit as st\n\n"
        f"GOLDEN_OCR_TEXT = {GOLDEN_OCR_TEXT!r}\n"
        f"GOLDEN_RECEIPT = "
        f"{pformat(GOLDEN_RECEIPT, sort_dicts=False, width=88)}\n"
        + "\n"
        + spatial_reconstruction_source().strip()
        + "\n\n"
        + parser_source().strip()
        + "\n\n"
        + dedent(
        f'''
        def read_receipt_now(uploaded):
            suffix = Path(uploaded.name).suffix.lower()
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temp:
                temp.write(uploaded.getvalue())
                path = temp.name
            try:
                from paddleocr import PaddleOCR
                engine = PaddleOCR(
                    lang="korean",
                    ocr_version="PP-OCRv5",
                    use_doc_orientation_classify=False,
                    use_doc_unwarping=False,
                    use_textline_orientation=False,
                    device="cpu",
                )
                page = list(engine.predict(path))[0]
                payload = page.json() if callable(page.json) else page.json
                result = payload.get("res", payload)
                items = [
                    {{
                        "page": 1,
                        "box": box.tolist() if hasattr(box, "tolist") else box,
                        "text": text,
                        "confidence": float(score),
                    }}
                    for box, text, score in zip(
                        result.get("rec_polys", []),
                        result.get("rec_texts", []),
                        result.get("rec_scores", []),
                    )
                ]
                return "\\n".join(reconstruct_spatial_lines(items))
            finally:
                Path(path).unlink(missing_ok=True)


        def process_document(uploaded=None, *, use_course_example=False):
            if use_course_example:
                text = GOLDEN_OCR_TEXT
                mode = "COURSE_EXAMPLE"
            elif uploaded is None:
                return {{"ok": False, "mode": "INPUT_ERROR", "error": "파일을 선택하세요."}}
            else:
                try:
                    text = read_receipt_now(uploaded)
                    mode = "DIRECT_OCR"
                except Exception as exc:
                    return {{
                        "ok": False,
                        "mode": "OCR_ERROR",
                        "error": f"{{type(exc).__name__}}: {{exc}}",
                        "recovery": "아래 ‘수업용 예제로 계속하기’를 누르세요.",
                    }}
            data = extract_receipt_from_text(
                text,
                "direct_ocr_rule_extraction" if mode == "DIRECT_OCR"
                else "course_example_rule_extraction",
            )
            return {{"ok": True, "mode": mode, "ocr_text": text, "data": data}}


        st.title("영수증 Document AI 연결 앱")
        uploaded = st.file_uploader(
            "승인된 비식별 이미지 또는 PDF 한 장 · 최대 5MB",
            type=["png", "jpg", "jpeg", "pdf"],
            max_upload_size=5,
            help="PNG, JPG, JPEG, PDF만 허용합니다. 수업에서는 한 번에 5MB 이하 한 장만 처리합니다.",
        )
        left, right = st.columns(2)
        run_ocr_now = left.button("내 영수증 직접 읽기", type="primary")
        use_example = right.button("수업용 예제로 계속하기")
        if run_ocr_now or use_example:
            result = process_document(
                uploaded,
                use_course_example=use_example,
            )
            if result["ok"]:
                if result["mode"] == "DIRECT_OCR":
                    st.success("지금 이 사진을 OCR로 직접 읽었습니다.")
                else:
                    st.info(
                        "수업용 예제 결과를 불러왔습니다. "
                        "현재 업로드한 파일을 분석한 결과가 아닙니다."
                    )
                st.text_area("OCR 원문", result["ocr_text"], height=220)
                st.json(result["data"])
            else:
                if result["mode"] == "INPUT_ERROR":
                    st.error("먼저 영수증 파일을 선택하세요.")
                else:
                    st.error("사진을 직접 읽지 못했습니다.")
                    st.caption(f"원인: {{result['error']}}")
                if result.get("recovery"):
                    st.info(result["recovery"])
        '''
        ).lstrip()
    )
    app_assignment = readable_string_assignment("app_code", app_source)
    cells = [
        intro(
            6,
            "OCR 및 정보 추출 기능 연동",
            "app_06.py",
            "업로드한 파일을 OCR 함수에 연결하고 직접 읽기·실패·수업용 예제를 화면에서 구분합니다.",
        ),
        runtime_cell(),
        code(STREAMLIT_SETUP + PADDLEOCR_SETUP),
        code(
            app_assignment
            + """

output_path = OUTPUT_DIR / "app_06.py"
output_path.write_text(app_code, encoding="utf-8")
print("저장:", output_path)
"""
        ),
        code(
            spatial_reconstruction_source()
            + "\n"
            + parser_source()
            + """

OCR_RECORD_PATH = "tests/fixtures/ppocrv5_recorded_receipt_tokens.json"
recorded_asset = load_course_assets(OCR_RECORD_PATH)
RECORDED_PP_OCRV5_TOKENS = json.loads(
    recorded_asset[OCR_RECORD_PATH].decode("utf-8")
)
recorded_text = "\\n".join(
    reconstruct_spatial_lines(RECORDED_PP_OCRV5_TOKENS)
)
recorded_receipt = extract_receipt_from_text(
    recorded_text,
    "recorded_ppocrv5_regression",
)
assert recorded_receipt["date"] == "2025-10-04"
assert recorded_receipt["total_amount"] == 76000
assert len(recorded_receipt["items"]) == 5
print(
    "실제 OCR 기록 재검사 통과:",
    recorded_receipt["total_amount"],
    len(recorded_receipt["items"]),
)
"""
        ),
        markdown(
            """
            ## 내가 직접 정하는 OCR 통과 조건 3개

            앱이 오류 없이 열리는 것과 추출값이 맞는 것은 다릅니다. 내 영수증을
            직접 읽은 뒤 반드시 확인해야 할 값을 세 개 고릅니다.
            """
        ),
        code(
            """
            # TODO: 확인할 필드 세 개를 채우세요.
            my_ocr_checks = [None, None, None]
            if any(value is None for value in my_ocr_checks):
                print("빈칸이 있습니다. 아래 전체 정답과 비교하세요.")
            """
        ),
        markdown(
            """
            <details>
            <summary>힌트와 전체 정답 보기</summary>

            날짜, 총액, 반복 품목 수는 후속 검증과 Excel에 직접 영향을 줍니다.
            </details>
            """
        ),
        code(
            """
            ANSWER_OCR_CHECKS = ["date", "total_amount", "items"]
            assert recorded_receipt["date"]
            assert recorded_receipt["total_amount"] is not None
            assert recorded_receipt["items"]
            print("전체 정답 · 직접 OCR 실행 후 필수 확인:", ANSWER_OCR_CHECKS)
            """
        ),
        code(
            """
            from streamlit.testing.v1 import AppTest

            app_test = AppTest.from_file(str(output_path)).run(timeout=20)
            assert not app_test.exception
            assert app_test.title[0].value == "영수증 Document AI 연결 앱"
            assert len(app_test.button) == 2
            app_test.button[1].click().run(timeout=20)
            assert any(
                "현재 업로드한 파일을 분석한 결과가 아닙니다" in item.value
                for item in app_test.info
            )
            assert app_test.json
            print("CHECKPOINT 1/1 PASS: 앱 연결·모드 표시·JSON 출력")
            """
        ),
        streamlit_preview_cell("output_path", 8506),
    ]
    return notebook("06_ocr_ai_integration.ipynb", cells)


def notebook_07() -> dict:
    cells = [
        intro(
            7,
            "추출 결과 검증 및 데이터 저장",
            "receipt_result.xlsx",
            "오류·경고·사람 검토를 분리하고, 공개된 승인 정답 경로에서만 Excel을 만듭니다.",
        ),
        runtime_cell(),
        code(STREAMLIT_SETUP),
        code(
            """
            import importlib.util
            import subprocess
            if importlib.util.find_spec("openpyxl") is None:
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", "-q", "openpyxl==3.1.5"]
                )
            from copy import deepcopy
            from datetime import date
            from openpyxl import Workbook, load_workbook
            """
        ),
        code(golden_constants()),
        code(
            """
            input_path = OUTPUT_DIR / "receipt.json"
            # 기본값 True: 새 Colab에서도 공개 준비 입력으로 바로 실행합니다.
            # 앞 교시 파일을 이어 쓰려면 False로 바꾸고 업로드 창에서 선택합니다.
            USE_COURSE_EXAMPLE = True
            if not input_path.exists() and not USE_COURSE_EXAMPLE:
                upload_previous_artifact("receipt.json")
            if input_path.exists():
                receipt = json.loads(input_path.read_text(encoding="utf-8"))
                INPUT_MODE = "PREVIOUS_LESSON"
            else:
                receipt = deepcopy(GOLDEN_RECEIPT)
                INPUT_MODE = "COURSE_EXAMPLE"


            def validate_receipt(data):
                warnings, errors = [], []
                for field in ("store_name", "date", "total_amount", "items"):
                    if data.get(field) in (None, "", []):
                        errors.append(f"필수값 누락: {field}")
                try:
                    parsed_date = date.fromisoformat(data.get("date", ""))
                    if parsed_date > date.today():
                        warnings.append("미래 날짜입니다. 원본을 확인하세요.")
                except ValueError:
                    errors.append("date는 YYYY-MM-DD 형식이어야 합니다.")

                total = data.get("total_amount")
                if isinstance(total, bool) or not isinstance(total, int) or total < 0:
                    errors.append("total_amount는 0 이상의 정수여야 합니다.")
                item_sum = 0
                for index, item in enumerate(data.get("items") or [], start=1):
                    values = [item.get(key) for key in ("quantity", "unit_price", "line_total")]
                    if not all(isinstance(value, int) and not isinstance(value, bool) for value in values):
                        errors.append(f"{index}번째 품목 금액 형식 오류")
                        continue
                    if values[0] * values[1] != values[2]:
                        errors.append(f"{index}번째 품목 수량×단가 오류")
                    item_sum += values[2]
                adjustments = data.get("adjustments") or {}
                expected = (
                    item_sum
                    - adjustments.get("discount", 0)
                    + adjustments.get("tax", 0)
                    + adjustments.get("service", 0)
                    + adjustments.get("rounding", 0)
                )
                if isinstance(total, int) and not isinstance(total, bool) and expected != total:
                    errors.append(f"품목·조정 후 합계 {expected:,}원과 총액 {total:,}원이 다릅니다.")
                tax_breakdown = data.get("tax_breakdown")
                if tax_breakdown and tax_breakdown.get("mode") == "included_in_item_prices":
                    supply = tax_breakdown.get("supply_amount")
                    vat = tax_breakdown.get("vat")
                    payable = tax_breakdown.get("payable_total")
                    if not all(isinstance(value, int) and not isinstance(value, bool)
                               for value in (supply, vat, payable)):
                        errors.append("포함세액 내역은 정수 금액이어야 합니다.")
                    elif supply + vat != payable or payable != total:
                        errors.append("공급가액·포함 부가세·총액 관계가 맞지 않습니다.")
                    if adjustments.get("tax", 0) != 0:
                        errors.append("포함 부가세를 adjustments.tax에 다시 더하면 이중 계산됩니다.")
                for field in ("store_name", "date", "total_amount"):
                    if not (data.get("evidence") or {}).get(field):
                        warnings.append(f"{field}의 원본 근거가 없습니다.")
                return {"valid": not errors, "warnings": warnings, "errors": errors}


            validation = validate_receipt(receipt)
            source_text = receipt.get("source_text") or GOLDEN_OCR_TEXT
            print(
                "입력 자료:",
                (
                    "4교시에서 만든 JSON을 불러왔습니다."
                    if INPUT_MODE == "PREVIOUS_LESSON"
                    else "수업용 예제 JSON을 불러왔습니다."
                ),
            )
            if INPUT_MODE == "COURSE_EXAMPLE":
                print("중요: 지금 업로드한 문서를 새로 분석한 결과가 아닙니다.")
            print("검증:", validation)
            if not validation["valid"]:
                print(
                    "BLOCKED_BY_VALIDATION: 아래 최종 앱에서 원본과 대조해 "
                    "값을 수정한 뒤 승인하세요."
                )
            """
        ),
        code(
            """
            def safe_text(value):
                if isinstance(value, str) and value.lstrip(" \\t\\r\\n").startswith(
                    ("=", "+", "-", "@")
                ):
                    return "'" + value
                return value


            def save_reviewed_excel(data, validation, review_record, output_path, source_text):
                if not validation["valid"]:
                    return False
                if review_record.get("decision") not in {"APPROVED", "CHANGED"}:
                    return False

                workbook = Workbook()
                summary = workbook.active
                summary.title = "검토_요약"
                summary.append([
                    "field", "raw_value", "cleaned_value", "final_value",
                    "decision", "reviewer", "reviewed_at", "change_reason",
                ])
                raw = data.get("raw_values") or {}
                cleaned = data.get("cleaned_values") or {}
                for field in ("store_name", "date", "total_amount"):
                    summary.append([
                        field,
                        safe_text(raw.get(field)),
                        safe_text(cleaned.get(field)),
                        safe_text(data.get(field)),
                        review_record["decision"],
                        safe_text(review_record["reviewer"]),
                        review_record["reviewed_at"],
                        safe_text(review_record["note"]),
                    ])

                items = workbook.create_sheet("품목")
                items.append(["품목", "수량", "단가", "금액"])
                for item in data["items"]:
                    items.append([
                        safe_text(item["name"]),
                        item["quantity"],
                        item["unit_price"],
                        item["line_total"],
                    ])

                evidence = workbook.create_sheet("원문_근거")
                evidence.append(["source_mode", data.get("source_mode")])
                evidence.append(["ocr_text", safe_text(source_text)])
                evidence.append(["evidence", safe_text(json.dumps(
                    data.get("evidence") or {}, ensure_ascii=False
                ))])
                workbook.save(output_path)
                return True
            """
        ),
        markdown(
            """
            ## 시나리오 A. 기본값은 차단

            사람이 원본을 보기 전에는 결과가 유효해도 다운로드를 열지 않습니다.
            """
        ),
        code(
            """
            blocked_path = OUTPUT_DIR / "pending_review.xlsx"
            PENDING_REVIEW = {
                "decision": "PENDING",
                "reviewer": "",
                "reviewed_at": "",
                "note": "",
            }
            assert not save_reviewed_excel(
                receipt, validation, PENDING_REVIEW, blocked_path, source_text
            )
            assert not blocked_path.exists()
            print("DEFAULT_BLOCKED PASS: 미승인 Excel 없음")
            """
        ),
        markdown(
            """
            ## 시나리오 B. 내가 직접 남기는 승인 기록

            원본의 상호명·날짜·품목·총액을 직접 대조한 뒤 세 곳을 채웁니다.
            """
        ),
        code(
            """
            # TODO: 원본 대조 뒤 세 곳을 채우세요.
            my_decision = None
            my_reviewer = None
            my_review_note = None
            if None in (my_decision, my_reviewer, my_review_note):
                print("빈칸이 있습니다. 아래 전체 정답과 비교하세요.")
            """
        ),
        markdown(
            """
            <details>
            <summary>힌트와 전체 정답 보기</summary>

            값 수정이 없으면 `APPROVED`, 수정했다면 `CHANGED`입니다. 검토자와
            무엇을 확인했는지도 기록합니다.
            </details>
            """
        ),
        code(
            """
            from datetime import datetime, timedelta, timezone

            KST = timezone(timedelta(hours=9))
            IS_COURSE_SAMPLE = (
                receipt.get("date") == "2025-10-04"
                and receipt.get("total_amount") == 76000
                and len(receipt.get("items") or []) == 5
            )
            reviewed_receipt = deepcopy(receipt)
            item_corrections = []
            if IS_COURSE_SAMPLE:
                for item, answer_item in zip(
                    reviewed_receipt["items"],
                    GOLDEN_RECEIPT["items"],
                ):
                    if item["name"] != answer_item["name"]:
                        item_corrections.append({
                            "OCR 판독": item["name"],
                            "원본 대조 후": answer_item["name"],
                        })
                        item["name"] = answer_item["name"]
                if item_corrections:
                    print("공개 샘플의 품목명 원본 대조 정답:")
                    for correction in item_corrections:
                        print(
                            "-",
                            correction["OCR 판독"],
                            "→",
                            correction["원본 대조 후"],
                        )

            learner_completed_review = None not in (
                my_decision,
                my_reviewer,
                my_review_note,
            )
            default_decision = "CHANGED" if item_corrections else "APPROVED"
            REVIEW_RECORD = {
                "decision": (
                    my_decision
                    if learner_completed_review
                    else default_decision if IS_COURSE_SAMPLE else "PENDING"
                ),
                "reviewer": (
                    my_reviewer
                    if learner_completed_review
                    else "공개 정답" if IS_COURSE_SAMPLE else ""
                ),
                "reviewed_at": datetime.now(KST).isoformat(timespec="seconds"),
                "note": (
                    my_review_note
                    if learner_completed_review
                    else (
                        "공개 비식별 원본 대조 후 OCR 품목명 수정"
                        if item_corrections
                        else "공개 비식별 원본과 주요 필드 대조 완료"
                    )
                    if IS_COURSE_SAMPLE
                    else ""
                ),
            }
            reviewed_validation = validate_receipt(reviewed_receipt)
            output_path = OUTPUT_DIR / "receipt_result.xlsx"
            excel_created = save_reviewed_excel(
                reviewed_receipt,
                reviewed_validation,
                REVIEW_RECORD,
                output_path,
                source_text,
            )
            if excel_created:
                saved = load_workbook(output_path)
                assert saved.sheetnames == ["검토_요약", "품목", "원문_근거"]
                assert saved["검토_요약"]["E2"].value in {"APPROVED", "CHANGED"}
                print("REVIEWED_APPROVED PASS:", output_path, saved.sheetnames)
                print("CHECKPOINT 1/1 PASS: 미승인 차단 + 승인 후 Excel")
                download_artifact(output_path)
            else:
                print(
                    "Excel 생성 차단: 검증 오류를 최종 앱에서 수정한 뒤 "
                    "다운로드하세요."
                )
            """
        ),
        markdown(
            """
            ## 최종 앱: 업로드부터 Excel 다운로드까지 한 화면으로 연결

            아래 셀은 앞 교시의 기능을 하나의 실행 가능한 앱으로 묶습니다.
            앱에서는 OCR/VLM 경로 선택, 원문·JSON 확인, 상호명·날짜·총액·품목
            수정, 재검증, 사람 승인, Excel 다운로드를 순서대로 수행합니다.
            """
        ),
        code(
            f"""
import shutil

FINAL_APP_SOURCE_PATHS = {FINAL_APP_SOURCE_PATHS!r}
final_app_assets = load_course_assets(*FINAL_APP_SOURCE_PATHS)

final_app_dir = OUTPUT_DIR / "final_document_ai_app"
for relative_path in FINAL_APP_SOURCE_PATHS:
    target = final_app_dir / relative_path
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(final_app_assets[relative_path])

final_app_path = final_app_dir / "app.py"
archive_base = OUTPUT_DIR / "final_document_ai_app"
archive_path = Path(
    shutil.make_archive(
        str(archive_base),
        "zip",
        root_dir=final_app_dir,
    )
)
print("최종 앱:", final_app_path)
print("앱 전체 코드:", archive_path)
"""
        ),
        code(
            """
            import sys
            from streamlit.testing.v1 import AppTest

            sys.path.insert(0, str(final_app_dir))
            final_test = AppTest.from_file(str(final_app_path)).run(timeout=30)
            assert not final_test.exception
            final_test.button(key="run_sample").click().run(timeout=30)
            assert not final_test.exception
            assert any(
                "원본 대조 후 수정" in item.value
                for item in final_test.subheader
            )
            assert len(final_test.get("download_button")) == 0
            final_test.checkbox(key="review_complete").check().run(timeout=30)
            assert not final_test.exception
            assert len(final_test.get("download_button")) == 1
            print("FINAL APP PASS: 수정 표·재검증·승인·Excel 다운로드")
            download_artifact(archive_path)
            """
        ),
        streamlit_preview_cell("final_app_path", 8507),
    ]
    return notebook("07_validation_export.ipynb", cells)


def notebook_08() -> dict:
    cells = [
        intro(
            8,
            "실무 적용 시나리오 설계 및 최종 정리",
            "poc_candidate_card.md",
            "견적서·신청서·거래명세서 실물 사진을 비교하고 첫 PoC 한 가지를 고릅니다.",
        ),
        runtime_cell(),
        code(
            f"""
            import io
            from PIL import Image
            try:
                from IPython.display import display
            except ImportError:
                display = lambda image: None

            EXTENSION_IMAGE_PATHS = {{
                "quotation": "sample_docs/extensions/quotation_photo.png",
                "application": "sample_docs/extensions/application_form_photo.png",
                "transaction_statement": (
                    "sample_docs/extensions/transaction_statement_photo.png"
                ),
            }}
            EXTENSION_EXAMPLES = {EXTENSION_EXAMPLES!r}
            extension_assets = load_course_assets(
                *EXTENSION_IMAGE_PATHS.values()
            )
            for key, path in EXTENSION_IMAGE_PATHS.items():
                image = Image.open(io.BytesIO(extension_assets[path])).convert("RGB")
                image.thumbnail((320, 400))
                print(key, image.size)
                display(image)
            """
        ),
        markdown(
            """
            ## 형식이 바뀌면 생기는 어려움

            - **Excel**: 수식, 병합 셀, 숨김 시트, 숫자 서식
            - **Word**: 머리글, 텍스트박스, 변경 추적, 이미지로 삽입된 본문
            - **PDF**: 텍스트·스캔 혼합 페이지, 암호, 깨진 문자맵
            - **PPT**: 그룹 도형, 읽기 순서, 발표자 노트
            - **표 캡처**: 셀 관계가 사라져 행·열 위상을 다시 복원해야 함
            """
        ),
        code(
            """
            import re
            import zipfile

            OFFICE_PATHS = [
                "sample_docs/formats/quotation.xlsx",
                "sample_docs/formats/application_form.docx",
                "sample_docs/formats/transaction_statement.pdf",
                "sample_docs/formats/table_summary.pptx",
            ]
            CODE_EXAMPLE_PATHS = [
                "src/document_examples.py",
                "sample_outputs/extensions/quotation.json",
                "sample_outputs/extensions/application.json",
                "sample_outputs/extensions/transaction_statement.json",
            ]
            all_assets = load_course_assets(
                *OFFICE_PATHS,
                *CODE_EXAMPLE_PATHS,
            )
            office_assets = {
                path: all_assets[path]
                for path in OFFICE_PATHS
            }
            office_dir = OUTPUT_DIR / "office_format_samples"
            office_dir.mkdir(exist_ok=True)
            for path, payload in office_assets.items():
                (office_dir / Path(path).name).write_bytes(payload)

            def xml_text_count(path, prefix, text_tag):
                with zipfile.ZipFile(path) as archive:
                    names = [
                        name for name in archive.namelist()
                        if name.startswith(prefix) and name.endswith(".xml")
                    ]
                    text_count = 0
                    for name in names:
                        xml = archive.read(name).decode("utf-8", errors="ignore")
                        text_count += len(re.findall(text_tag, xml))
                    return len(names), text_count

            xlsx_sheets, xlsx_values = xml_text_count(
                office_dir / "quotation.xlsx",
                "xl/worksheets/",
                r"<x:(?:v|f)>",
            )
            docx_parts, docx_text = xml_text_count(
                office_dir / "application_form.docx",
                "word/document",
                r"<w:t",
            )
            pptx_slides, pptx_text = xml_text_count(
                office_dir / "table_summary.pptx",
                "ppt/slides/slide",
                r"<a:t>",
            )
            pdf_bytes = (office_dir / "transaction_statement.pdf").read_bytes()
            print("Excel:", xlsx_sheets, "개 시트 XML · 값/수식", xlsx_values)
            print("Word:", docx_text, "개 본문 텍스트 run · 이미지 본문 여부 확인")
            print("PDF:", pdf_bytes[:5], "· 텍스트층 샘플")
            print("PPT:", pptx_slides, "개 슬라이드 · 텍스트", pptx_text)

            office_bundle = OUTPUT_DIR / "office_format_samples.zip"
            with zipfile.ZipFile(office_bundle, "w") as archive:
                for path in sorted(office_dir.iterdir()):
                    archive.write(path, path.name)
            print("실제 파일 4종 묶음:", office_bundle)
            download_artifact(office_bundle)

            code_example_dir = OUTPUT_DIR / "business_document_code_examples"
            code_example_dir.mkdir(exist_ok=True)
            for path in CODE_EXAMPLE_PATHS:
                target = code_example_dir / Path(path).name
                target.write_bytes(all_assets[path])
                if target.suffix == ".json":
                    payload = json.loads(all_assets[path].decode("utf-8"))
                    print(
                        "확장 JSON:",
                        payload["document_type"],
                        "· 필드",
                        len(payload),
                    )

            code_example_bundle = (
                OUTPUT_DIR / "business_document_code_examples.zip"
            )
            with zipfile.ZipFile(code_example_bundle, "w") as archive:
                for path in sorted(code_example_dir.iterdir()):
                    archive.write(path, path.name)
            print("문서별 Python·JSON 예제:", code_example_bundle)
            download_artifact(code_example_bundle)
            """
        ),
        markdown(
            """
            ## 내가 직접 만드는 PoC 카드

            `candidate`는 `quotation`, `application`, `transaction_statement`
            중 하나입니다. 점수는 1~5점이며 오류 영향과 예외 빈도는 낮을수록
            첫 PoC에 유리합니다.
            """
        ),
        code(
            """
            # TODO: 내 업무 후보와 점수·검토자·중단 조건을 채우세요.
            candidate = None
            score = {
                "반복량": None,
                "필드 안정성": None,
                "오류 영향": None,
                "예외 빈도": None,
                "사람 검토 가능성": None,
            }
            review_owner = None
            stop_condition = None
            if candidate is None or any(value is None for value in score.values()):
                print("빈칸이 있습니다. 아래 힌트·전체 정답과 비교하세요.")
            """
        ),
        markdown(
            """
            <details>
            <summary>힌트와 전체 정답 보기</summary>

            예시는 거래명세서를 한 장씩 처리하고 정산 담당자가 검토하는 작은
            PoC입니다. 값이 맞지 않거나 원본 근거가 없으면 저장을 중단합니다.
            </details>
            """
        ),
        code(
            """
            from textwrap import dedent

            candidate = candidate or "transaction_statement"
            if candidate not in EXTENSION_EXAMPLES:
                raise ValueError(
                    "candidate는 quotation, application, "
                    "transaction_statement 중 하나여야 합니다."
                )
            default_score = {
                "반복량": 4,
                "필드 안정성": 4,
                "오류 영향": 2,
                "예외 빈도": 3,
                "사람 검토 가능성": 5,
            }
            score = {
                key: (
                    int(value)
                    if value is not None
                    else default_score[key]
                )
                for key, value in score.items()
            }
            if not all(1 <= value <= 5 for value in score.values()):
                raise ValueError("모든 점수는 1~5 사이여야 합니다.")
            review_owner = review_owner or "정산 담당자"
            stop_condition = (
                stop_condition
                or "필수값·합계·원본 근거 중 하나라도 틀리면 자동 저장 중단"
            )
            example = EXTENSION_EXAMPLES[candidate]
            recommendation = (
                "GO_SMALL"
                if (
                    score["반복량"] >= 4
                    and score["필드 안정성"] >= 3
                    and score["오류 영향"] <= 3
                    and score["예외 빈도"] <= 3
                    and score["사람 검토 가능성"] >= 4
                )
                else "REVIEW"
            )
            card = f'''# 문서 자동화 PoC 후보 카드

            | 항목 | 내용 |
            | --- | --- |
            | 선택 문서 | {example["name"]} |
            | 추출 필드 | {", ".join(example["fields"])} |
            | 검증 규칙 | {" / ".join(example["rules"])} |
            | 틀렸을 때 영향 | {example["risk"]} |
            | 입력 제한 | 승인된 비식별 한 장 |
            | 최종 산출물 | 사람 승인 후 Excel |
            | 사람 검토자 | {review_owner} |
            | 중단 조건 | {stop_condition} |
            | 점수 | {" / ".join(f"{key} {value}" for key, value in score.items())} |
            | 제안 | {recommendation} |

            ## 첫 PoC 통과 기준

            - 같은 양식 30장을 모아 정답표와 비교한다.
            - 필드별 정확도뿐 아니라 수정률과 처리시간을 기록한다.
            - 오류 시 자동 저장하지 않고 검토 대기열로 보낸다.
            - 개인정보·보존·삭제 정책을 먼저 승인받는다.
            '''
            output_path = OUTPUT_DIR / "poc_candidate_card.md"
            output_path.write_text(dedent(card), encoding="utf-8")
            print(dedent(card))
            print("CHECKPOINT 1/1 PASS:", output_path)
            download_artifact(output_path)
            """
        ),
    ]
    return notebook("08_business_application.ipynb", cells)


BUILDERS = {
    1: notebook_01,
    2: notebook_02,
    3: notebook_03,
    4: notebook_04,
    5: notebook_05,
    6: notebook_06,
    7: notebook_07,
    8: notebook_08,
}


def main() -> None:
    COLAB_DIR.mkdir(parents=True, exist_ok=True)
    for lesson, builder in BUILDERS.items():
        path = COLAB_DIR / f"{lesson:02d}_{NOTEBOOK_SLUGS[lesson]}.ipynb"
        path.write_text(
            json.dumps(builder(), ensure_ascii=False, indent=1) + "\n",
            encoding="utf-8",
        )
        print("생성:", path.relative_to(ROOT))


if __name__ == "__main__":
    main()
