"""Colab 노트북의 준비 경로와 공개된 승인 시나리오를 실행 검증한다."""

from __future__ import annotations

import json
import os
import re
import tempfile
import zipfile
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
COLAB_DIR = ROOT / "colab"

EXPECTED_ARTIFACTS = {
    "01_document_ai_overview.ipynb": "lesson01_comparison_report.json",
    "02_ocr_basic.ipynb": "ocr_result.json",
    "03_document_structure.ipynb": "clean_receipt.json",
    "04_genai_extraction.ipynb": "receipt.json",
    "05_streamlit_basic.ipynb": "app_05.py",
    "06_ocr_ai_integration.ipynb": "app_06.py",
    "07_validation_export.ipynb": "receipt_result.xlsx",
    "08_business_application.ipynb": "poc_candidate_card.md",
}


def validate_structure(path: Path, notebook: dict) -> None:
    assert notebook["nbformat"] == 4, path
    assert notebook["metadata"]["colab"]["name"] == path.name, path
    ids = [cell["id"] for cell in notebook["cells"]]
    assert len(ids) == len(set(ids)), f"{path}: duplicate cell ids"
    code_cells = [cell for cell in notebook["cells"] if cell["cell_type"] == "code"]
    assert code_cells, f"{path}: no code cells"
    assert all(cell["outputs"] == [] for cell in code_cells), path
    learning_steps = [
        cell["metadata"].get("learning_step")
        for cell in code_cells
    ]
    assert all(learning_steps), f"{path}: missing learning step metadata"
    assert [step["current"] for step in learning_steps] == list(
        range(1, len(code_cells) + 1)
    ), f"{path}: invalid learning step order"
    assert all(step["total"] == len(code_cells) for step in learning_steps), path
    assert all(
        step["title"]
        and step["action"]
        and step["expected"]
        and step["code_help"]
        and isinstance(step["learner_edits"], bool)
        and step["edit_kind"] in {"required", "optional", "none"}
        for step in learning_steps
    ), f"{path}: incomplete learning step guide"
    assert all("show_lab_step(" in cell["source"] for cell in code_cells), path
    assert all("complete_lab_step(" in cell["source"] for cell in code_cells), path
    assert all("# ── 코드 읽기" in cell["source"] for cell in code_cells), path

    source = "\n".join(cell["source"] for cell in notebook["cells"])
    for cell, step in zip(code_cells, learning_steps, strict=True):
        mentioned_identifiers = re.findall(
            r"`([A-Za-z_][A-Za-z0-9_]*(?:\(\))?)`",
            step["code_help"],
        )
        for mention in mentioned_identifiers:
            identifier = mention.removesuffix("()")
            assert identifier in cell["source"], (
                f"{path}: code_help mentions {mention} outside its code cell"
            )
    # 생성 앱 전체 소스처럼 긴 문자열 안의 임의 문자열은 제외하고,
    # 사람이 읽는 짧은 코드 줄에서 도구명 잔존 여부를 검사한다.
    visible_source = "\n".join(
        line for line in source.splitlines() if len(line) < 500
    )
    assert "Google Colab도 외부 클라우드" in source, path
    assert "## 코드 셀을 읽는 방법" in source, path
    assert "✅ 실습 완료" in source, path
    assert "TODO" in source, f"{path}: 학습자 빈칸이 없습니다."
    assert "전체 정답" in source, f"{path}: 공개 정답이 없습니다."
    assert "RESEARCH_NOTE_CELL" in source, path
    assert "다른 자료 실험 기록" in source, path
    assert code_cells[0]["metadata"].get("cellView") == "form", path
    assert code_cells[-1]["metadata"].get("cellView") == "form", path
    for cell in code_cells:
        step = cell["metadata"]["learning_step"]
        if (
            step["edit_kind"] == "none"
            and len(cell["source"].splitlines()) >= 45
        ):
            assert cell["metadata"].get("cellView") == "form", (
                f"{path}: 긴 읽기 전용 코드가 펼쳐져 있습니다: {step['title']}"
            )
    assert "easyocr" not in source.lower(), path
    assert "gradio" not in source.lower(), path
    assert "codex" not in visible_source.lower(), path
    for confusing_term in (
        "LIVE",
        "COURSE_EXAMPLE",
        "PROCESSING_PATH",
        "RUN_OCR_NOW",
        "USE_MY_FILE",
        "USE_COURSE_EXAMPLE",
        "INPUT_MODE",
        "PREVIOUS_LESSON",
        "GOLDEN_",
        "VALIDATION_MODE",
        "CHECKPOINT",
        "DEFAULT_BLOCKED",
        "REVIEWED_APPROVED",
        "FINAL APP PASS",
        "GO_SMALL",
        "BLOCKED_BY_VALIDATION",
        "APPROVED",
        "PENDING_REVIEW",
        "DOCUMENT_SOURCE",
        "INPUT_CHOICE",
        "RESULT_JSON_URL",
        "MATERIAL_SOURCE",
        "WHAT_WORKED",
        "WHAT_FAILED",
        "NEXT_QUESTION",
    ):
        assert confusing_term.lower() not in source.lower(), (
            f"{path}: confusing learner term remains: {confusing_term}"
        )
    assert "base64.b64decode" not in source, path
    assert "FALLBACK_IMAGE_BASE64" not in source, path
    assert "GOLDEN_IMAGE_BASE64" not in source, path
    longest_code_line = max(
        len(line)
        for cell in code_cells
        for line in cell["source"].splitlines()
    )
    assert longest_code_line < 500, (
        f"{path}: unreadable code line ({longest_code_line} chars)"
    )
    longest_code_cell = max(
        len(cell["source"].splitlines())
        for cell in code_cells
    )
    assert longest_code_cell < 400, (
        f"{path}: oversized code cell ({longest_code_cell} lines)"
    )
    if path.name == "01_document_ai_overview.ipynb":
        assert len(notebook["cells"]) >= 20
        assert "RECORDED_PP_OCRV5_TOKENS" in source
        assert "VLM_DRAFT_WITH_ERROR" in source
        assert "MY_OCR_REVIEW" in source
        assert "MY_CORRECTED_TOTAL" in source
        assert "validate_candidate" in source
        assert "MY_CONCEPT_CHOICES" in source
        assert "멀티모달 AI" in source
        assert "learner_attempts" in source
        assert source.count("# TODO") >= 6
        assert "16000" in source and "76000" in source
        assert "my_role_map" not in source
        assert "ANSWER_ROLE_MAP" not in source
        assert "receipt_pipeline_trace" not in source
        assert "taebaek_restaurant_2025_redacted.png" in source
        assert "ppocrv5_recorded_receipt_tokens.json" in source
        assert "ppocrv5_recorded_receipt_metadata.json" in source
        assert "files.upload()" in source
        assert 'OCR_RECORD_METADATA["coordinate_space"]["height"]' in source
        assert "expected_coordinate_height" in source
        assert "point[0] * scale_x" in source
        assert "point[1] * scale_y" in source
    if path.name == "02_ocr_basic.ipynb":
        assert "READ_CURRENT_IMAGE = not AUTOMATED_CHECK" in source
        assert 'lang="korean"' in source
        assert 'ocr_version="PP-OCRv5"' in source
        assert "ppocrv5_recorded_receipt_tokens.json" in source
        assert "ppocrv5_recorded_receipt_metadata.json" in source
        assert '실습_자료 = "제공 예제"' in source
        assert '"내 컴퓨터에서 업로드"' in source
        assert '"인터넷 이미지 URL"' in source
        assert "인터넷_이미지_URL" in source
        assert "OCR_COORDINATE_SIZE = receipt_image.size" in source
        assert "DISPLAY_INPUT_FILE_NAME" in source
        assert "인식한 글자와 신뢰도" in source
        assert "scaled_points + [scaled_points[0]]" in source
        assert "draw.rectangle(" not in source
    if path.name in {
        "03_document_structure.ipynb",
        "04_genai_extraction.ipynb",
        "07_validation_export.ipynb",
    }:
        assert '실습_자료 = "제공 예제"' in source
        assert '"인터넷 JSON URL"' in source
        assert "인터넷_JSON_URL" in source
    if path.name == "04_genai_extraction.ipynb":
        assert "evidence" in source and "provenance" in source
        assert '"engine": "이 노트북에서는 실행하지 않음"' in source
        assert "현재 실행에서 VLM을 호출한 결과가 아닙니다." in source
    if path.name == "05_streamlit_basic.ipynb":
        assert "uploaded.getvalue()" in source
        assert "serve_kernel_port_as_iframe" in source
    if path.name == "06_ocr_ai_integration.ipynb":
        assert "read_receipt_now" in source
        assert "실제 OCR 기록 재검사 통과" in source
        assert "serve_kernel_port_as_iframe" in source
        assert "현재 파일 직접 처리" in source
        assert "OCR 실행 실패" in source
        assert "내 영수증 직접 읽기" in source
        assert "수업용 예제로 계속하기" in source
        assert "현재 업로드한 파일을 분석한 결과가 아닙니다" in source
        assert "finally:" in source and "unlink(missing_ok=True)" in source
        assert "ppocrv5_recorded_receipt_tokens.json" in source
    if path.name == "07_validation_export.ipynb":
        assert "승인 전 저장 차단 확인" in source
        assert "승인 후 Excel 생성 확인" in source
        assert "최종 앱 자동검사 통과" in source
        assert "final_document_ai_app" in source
        assert "make_archive" in source
        assert "st.data_editor" in (ROOT / "app.py").read_text(encoding="utf-8")
        assert "serve_kernel_port_as_iframe" in source
        assert "REVIEW_RECORD" in source
        assert "검토_요약\", \"품목\", \"원문_근거" in source
        assert "FINAL_APP_SOURCE_PATHS" in source
        assert "final_app_assets = load_course_assets" in source
    if path.name == "08_business_application.ipynb":
        for document in ("quotation", "application", "transaction_statement"):
            assert document in source
        assert "office_format_samples.zip" in source
        assert "candidate = None" in source
        assert "score = {" in source
        assert "sample_docs/formats/quotation.xlsx" in source
        assert "sample_docs/extensions/quotation_photo.png" in source
        assert "business_document_code_examples.zip" in source
        assert "src/document_examples.py" in source


def execute_course_example_path(path: Path, notebook: dict) -> None:
    namespace = {"__name__": "__notebook_validation__"}
    with tempfile.TemporaryDirectory(prefix=f"{path.stem}_") as temp_dir:
        previous = Path.cwd()
        previous_flag = os.environ.get("COURSE_VALIDATE_EXAMPLE")
        previous_asset_root = os.environ.get("COURSE_LOCAL_ASSET_ROOT")
        os.environ["COURSE_VALIDATE_EXAMPLE"] = "1"
        os.environ["COURSE_LOCAL_ASSET_ROOT"] = str(ROOT)
        os.chdir(temp_dir)
        try:
            for index, cell in enumerate(notebook["cells"], start=1):
                if cell["cell_type"] != "code":
                    continue
                try:
                    exec(
                        compile(cell["source"], f"{path.name}:cell-{index}", "exec"),
                        namespace,
                    )
                except Exception as exc:
                    raise RuntimeError(
                        f"{path.name}의 {index}번째 셀 실행 실패"
                    ) from exc

            artifact = Path("course_outputs") / EXPECTED_ARTIFACTS[path.name]
            assert artifact.is_file(), f"{path}: missing {artifact.name}"
            assert artifact.stat().st_size > 0, f"{path}: empty {artifact.name}"
            if path.name == "01_document_ai_overview.ipynb":
                payload = json.loads(artifact.read_text(encoding="utf-8"))
                assert payload["ocr"]["token_count"] == 44
                assert payload["vlm"]["model_called_in_this_notebook"] is False
                assert payload["document_ai"]["before_validation"]["valid"] is False
                assert payload["document_ai"]["after_validation"]["valid"] is True
                assert payload["document_ai"]["corrected_total"] == 76000
                assert payload["idp"]["human_review_decision"] == "승인"
                assert len(payload["concept_answers"]) == 5
                assert len(payload["learner_attempts"]) == 8
                tokens = namespace["RECORDED_PP_OCRV5_TOKENS"]
                image = namespace["receipt_image"]
                scale_x = namespace["scale_x"]
                scale_y = namespace["scale_y"]
                scaled_x = [
                    point[0] * scale_x
                    for token in tokens
                    for point in token["box"]
                ]
                scaled_y = [
                    point[1] * scale_y
                    for token in tokens
                    for point in token["box"]
                ]
                assert max(scaled_x) > image.width * 0.9
                assert max(scaled_y) > image.height * 0.85
                assert abs(scale_x - scale_y) < 0.01
                assert namespace["OCR_COORDINATE_SIZE"] == (900, 1003)
            if path.name == "02_ocr_basic.ipynb":
                payload = json.loads(artifact.read_text(encoding="utf-8"))
                assert payload["result_source"] == "제공 예제 사용"
                assert payload["image_size"] == {
                    "width": 2558,
                    "height": 2850,
                }
                assert payload["ocr_coordinate_size"] == {
                    "width": 900,
                    "height": 1003,
                }
                assert abs(namespace["scale_x"] - namespace["scale_y"]) < 0.01
            research_note = Path(
                f"course_outputs/lesson{path.name[:2]}_research_note.md"
            )
            assert research_note.is_file(), f"{path}: missing research note"
            if path.name == "07_validation_export.ipynb":
                assert not Path("course_outputs/blocked_before_review.xlsx").exists()
                assert namespace["review_not_started"]["decision"] == "검토 전"
                assert namespace["REVIEW_RECORD"]["decision"] == "승인"
                with zipfile.ZipFile(
                    "course_outputs/final_document_ai_app.zip"
                ) as archive:
                    packaged_files = {
                        name for name in archive.namelist()
                        if not name.endswith("/")
                    }
                    assert packaged_files == set(
                        namespace["FINAL_APP_SOURCE_PATHS"]
                    )
            if path.name == "08_business_application.ipynb":
                with zipfile.ZipFile(
                    "course_outputs/business_document_code_examples.zip"
                ) as archive:
                    assert set(archive.namelist()) == {
                        "application.json",
                        "document_examples.py",
                        "quotation.json",
                        "transaction_statement.json",
                    }
        finally:
            os.chdir(previous)
            if previous_flag is None:
                os.environ.pop("COURSE_VALIDATE_EXAMPLE", None)
            else:
                os.environ["COURSE_VALIDATE_EXAMPLE"] = previous_flag
            if previous_asset_root is None:
                os.environ.pop("COURSE_LOCAL_ASSET_ROOT", None)
            else:
                os.environ["COURSE_LOCAL_ASSET_ROOT"] = previous_asset_root


def execute_sequential_handoff(paths: list[Path]) -> None:
    """02→03→04→07을 같은 산출물 폴더에서 새 노트북처럼 이어 실행한다."""

    selected = [
        path
        for lesson in ("02_", "03_", "04_", "07_")
        for path in paths
        if path.name.startswith(lesson)
    ]
    with tempfile.TemporaryDirectory(prefix="sequential_handoff_") as temp_dir:
        previous = Path.cwd()
        previous_flag = os.environ.get("COURSE_VALIDATE_EXAMPLE")
        previous_asset_root = os.environ.get("COURSE_LOCAL_ASSET_ROOT")
        os.environ["COURSE_VALIDATE_EXAMPLE"] = "1"
        os.environ["COURSE_LOCAL_ASSET_ROOT"] = str(ROOT)
        os.chdir(temp_dir)
        try:
            for path in selected:
                notebook = json.loads(path.read_text(encoding="utf-8"))
                namespace = {"__name__": "__sequential_validation__"}
                for index, cell in enumerate(notebook["cells"], start=1):
                    if cell["cell_type"] != "code":
                        continue
                    try:
                        exec(
                            compile(
                                cell["source"],
                                f"{path.name}:sequential-cell-{index}",
                                "exec",
                            ),
                            namespace,
                        )
                    except Exception as exc:
                        raise RuntimeError(
                            f"순차 handoff에서 {path.name}의 {index}번째 셀 실행 실패"
                        ) from exc

            output_dir = Path("course_outputs")
            ocr_payload = json.loads(
                (output_dir / "ocr_result.json").read_text(encoding="utf-8")
            )
            assert len(ocr_payload["items"]) >= 10
            clean_payload = json.loads(
                (output_dir / "clean_receipt.json").read_text(encoding="utf-8")
            )
            assert len(clean_payload["groups"]["items"]) == 5
            receipt_payload = json.loads(
                (output_dir / "receipt.json").read_text(encoding="utf-8")
            )
            assert len(receipt_payload["items"]) == 5
            assert receipt_payload["items"][2]["name"] == "수제돈가스"
            assert (output_dir / "receipt_result.xlsx").is_file()
            workbook = load_workbook(
                output_dir / "receipt_result.xlsx",
                read_only=True,
            )
            exported_item_names = [
                row[0]
                for row in workbook["품목"].iter_rows(
                    min_row=2,
                    values_only=True,
                )
            ]
            assert "수제 돈가스" in exported_item_names
            assert "수제돈가스" not in exported_item_names
        finally:
            os.chdir(previous)
            if previous_flag is None:
                os.environ.pop("COURSE_VALIDATE_EXAMPLE", None)
            else:
                os.environ["COURSE_VALIDATE_EXAMPLE"] = previous_flag
            if previous_asset_root is None:
                os.environ.pop("COURSE_LOCAL_ASSET_ROOT", None)
            else:
                os.environ["COURSE_LOCAL_ASSET_ROOT"] = previous_asset_root


def main() -> None:
    paths = sorted(COLAB_DIR.glob("*.ipynb"))
    assert len(paths) == 8, f"expected 8 notebooks, found {len(paths)}"
    for path in paths:
        notebook = json.loads(path.read_text(encoding="utf-8"))
        validate_structure(path, notebook)
        execute_course_example_path(path, notebook)
        print("OK:", path.name)
    execute_sequential_handoff(paths)
    print("OK: 02→03→04→07 동일 폴더 순차 handoff")
    print(
        f"검증 완료: {len(paths)}개 노트북 · 독립 복구 경로 · "
        "공개 승인 경로 · 순차 handoff"
    )


if __name__ == "__main__":
    main()
