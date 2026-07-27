"""8교시 Office 형식 체험 파일과 실물형 문서 사진을 구조적으로 검증한다."""

from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

import fitz
from openpyxl import load_workbook
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
FORMATS = ROOT / "sample_docs" / "formats"
EXTENSIONS = ROOT / "sample_docs" / "extensions"


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def verify_xlsx() -> None:
    path = FORMATS / "quotation.xlsx"
    workbook = load_workbook(path, data_only=False)
    require(
        workbook.sheetnames == ["견적_요약", "품목", "형식_주의사항"],
        "견적서 Excel의 시트 구성이 다릅니다.",
    )
    summary = workbook["견적_요약"]
    require(summary["B4"].value == "한빛오피스", "견적서 공급자가 다릅니다.")
    require(summary["D4"].value == "=SUM('품목'!E2:E3)", "공급가액 수식이 없습니다.")
    require(summary["F4"].value == "=D4+E4", "총액 수식이 없습니다.")
    print("OK: quotation.xlsx · 3시트 · 계산 수식")


def xml_text(path: Path, prefix: str) -> str:
    with ZipFile(path) as archive:
        names = archive.namelist()
        selected = [name for name in names if name.startswith(prefix) and name.endswith(".xml")]
        return "\n".join(archive.read(name).decode("utf-8") for name in selected)


def verify_docx() -> None:
    path = FORMATS / "application_form.docx"
    with ZipFile(path) as archive:
        names = archive.namelist()
        require("word/document.xml" in names, "Word 본문이 없습니다.")
        require(
            any(name.startswith("word/media/") for name in names),
            "이미지 기반 Word 양식의 본문 이미지가 없습니다.",
        )
        document_xml = archive.read("word/document.xml").decode("utf-8")
        require("a:blip" in document_xml, "Word 본문 이미지 연결이 없습니다.")
    print("OK: application_form.docx · 이미지 기반 Word 양식")


def verify_pdf() -> None:
    path = FORMATS / "transaction_statement.pdf"
    document = fitz.open(path)
    require(document.page_count == 1, "거래명세서 PDF는 한 장이어야 합니다.")
    text = document[0].get_text()
    for token in ("거 래 명 세 서", "다온유통", "91,300원"):
        require(token in text, f"거래명세서 PDF에서 {token!r}을 찾지 못했습니다.")
    print("OK: transaction_statement.pdf · 1페이지 · 텍스트층")


def verify_pptx() -> None:
    path = FORMATS / "table_summary.pptx"
    text = xml_text(path, "ppt/slides/slide")
    with ZipFile(path) as archive:
        slides = [
            name
            for name in archive.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        ]
    require(len(slides) == 1, "PPT 형식 체험 파일은 한 슬라이드여야 합니다.")
    for token in ("표 캡처는", "품목 행의 시작과 끝", "세액·총액", "원본 위치"):
        require(token in text, f"PPT에서 {token!r}을 찾지 못했습니다.")
    print("OK: table_summary.pptx · 1슬라이드 · 핵심 메시지")


def verify_document_photos() -> None:
    expected = [
        EXTENSIONS / "quotation_photo.png",
        EXTENSIONS / "application_form_photo.png",
        EXTENSIONS / "transaction_statement_photo.png",
    ]
    for path in expected:
        with Image.open(path) as image:
            require(image.width >= 1000 and image.height >= 1200, f"{path.name} 해상도가 낮습니다.")
            require(image.mode in {"RGB", "RGBA"}, f"{path.name} 색상 모드가 예상과 다릅니다.")
        print(f"OK: {path.name} · 실물형 합성 사진")


def main() -> None:
    verify_xlsx()
    verify_docx()
    verify_pdf()
    verify_pptx()
    verify_document_photos()
    print("검증 완료: Office 4종 · 실물형 문서 사진 3종")


if __name__ == "__main__":
    main()
