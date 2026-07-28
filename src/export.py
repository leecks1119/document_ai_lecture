"""7교시: 검증된 영수증 JSON을 안전한 Excel 파일로 바꾼다."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


FORMULA_PREFIXES = ("=", "+", "-", "@")


def safe_spreadsheet_text(value: Any) -> Any:
    """스프레드시트에서 수식으로 해석될 수 있는 문자열을 보호한다."""

    if isinstance(value, str) and value.lstrip(" \t\r\n").startswith(
        FORMULA_PREFIXES
    ):
        return "'" + value
    return value


def receipt_to_rows(data: dict) -> list[dict]:
    """영수증의 반복 품목을 Excel의 여러 행으로 펼친다."""

    base = {
        "store_name": data.get("store_name"),
        "date": data.get("date"),
        "total_amount": data.get("total_amount"),
    }
    rows = []
    for item in data.get("items") or []:
        row = {
            **base,
            "item_name": item.get("name"),
            "quantity": item.get("quantity"),
            "unit_price": item.get("unit_price"),
            "line_total": item.get("line_total"),
        }
        rows.append(
            {key: safe_spreadsheet_text(value) for key, value in row.items()}
        )
    return rows


def _write_table(sheet, columns: list[str], rows: list[dict]) -> None:
    """작은 표를 쓰고 초보자용 기본 서식을 적용한다."""

    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="173B57")

    for row in rows:
        sheet.append([safe_spreadsheet_text(row.get(column)) for column in columns])

    sheet.freeze_panes = "A2"
    for index, column in enumerate(columns, start=1):
        values = [str(column)]
        values.extend(
            str(sheet.cell(row=row, column=index).value or "")
            for row in range(2, sheet.max_row + 1)
        )
        width = min(max(len(value) for value in values) + 2, 40)
        sheet.column_dimensions[get_column_letter(index)].width = width


def receipt_to_xlsx_bytes(
    data: dict,
    *,
    source_text: str = "",
    review_status: str = "APPROVED",
    review_record: dict | None = None,
) -> bytes:
    """원문·정제값·최종값·검토 상태가 남는 Excel 바이트를 반환한다."""

    workbook = Workbook()
    summary = workbook.active
    summary.title = "검토_요약"

    raw_values = data.get("raw_values") or {}
    cleaned_values = data.get("cleaned_values") or {}
    review_record = review_record or {
        "decision": review_status,
        "reviewer": "교육용 검수자",
        "reviewed_at": "",
        "note": "",
    }
    fields = ("store_name", "date", "total_amount")
    summary_rows = [
        {
            "field": field,
            "raw_value": raw_values.get(field, data.get(field)),
            "cleaned_value": cleaned_values.get(field, data.get(field)),
            "final_value": data.get(field),
            "decision": review_record.get("decision", review_status),
            "reviewer": review_record.get("reviewer", ""),
            "reviewed_at": review_record.get("reviewed_at", ""),
            "change_reason": review_record.get("note", ""),
        }
        for field in fields
    ]
    _write_table(
        summary,
        [
            "field",
            "raw_value",
            "cleaned_value",
            "final_value",
            "decision",
            "reviewer",
            "reviewed_at",
            "change_reason",
        ],
        summary_rows,
    )

    items = workbook.create_sheet("품목")
    item_columns = [
        "store_name",
        "date",
        "total_amount",
        "item_name",
        "quantity",
        "unit_price",
        "line_total",
    ]
    _write_table(items, item_columns, receipt_to_rows(data))

    source = workbook.create_sheet("원문_근거")
    source.append(["source_mode", safe_spreadsheet_text(data.get("source_mode", ""))])
    provenance = data.get("provenance") or {}
    for key in (
        "fixture_type",
        "input_file",
        "input_sha256",
        "engine",
        "engine_version",
        "target_technology",
        "recorded_at",
        "reviewer",
        "disclaimer",
    ):
        source.append([key, safe_spreadsheet_text(provenance.get(key, ""))])
    source.append(["ocr_text", safe_spreadsheet_text(source_text)])
    source.append(["evidence", safe_spreadsheet_text(str(data.get("evidence", {})))])
    source.column_dimensions["A"].width = 18
    source.column_dimensions["B"].width = 80

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
